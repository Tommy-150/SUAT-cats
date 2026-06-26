#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SUAT-cats 管理器
================
行列表视图，支持上移/下移步数调整顺序，内建预览服务器。
集成：基本信息编辑、故事编辑、图片预览、缩略图编辑、新增猫咪、删除猫咪、自动同步前端（Excel + cats.json）。
关闭时若有未保存的顺序变更会提醒。
依赖：Pillow、openpyxl

修改说明（支持编辑已有猫咪的图名）：
- CatEditor._build：移除已有猫咪图名输入框的 disabled 状态，调整提示文字。
- CatEditor._save：编辑已有猫咪时，若图名变化，先在旧文件夹内重命名所有图片文件，
  失败则回滚并阻止保存，成功后继续原有流程。
"""
import os
import re
import sys
import json
import glob
import shutil
import traceback
import subprocess
import threading
import webbrowser
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.font as tkfont

try:
    from PIL import Image, ImageTk
except ImportError:
    print("缺少 Pillow：pip install Pillow")
    sys.exit(1)
try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少 openpyxl：pip install openpyxl")
    sys.exit(1)


# ============================================================
# Windows 高 DPI 适配
# ============================================================
def _enable_dpi_awareness():
    if sys.platform != "win32":
        return
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
    except Exception:
        pass


_enable_dpi_awareness()


# ============================================================
# 视觉常量
# ============================================================
COLOR_BG = "#F7F5F0"
COLOR_CARD = "#FFFFFF"
COLOR_CARD_BORDER = "#ECE7DC"
COLOR_TEXT = "#333333"
COLOR_SUBTEXT = "#888888"
COLOR_ACCENT = "#4A4A4A"
COLOR_ACCENT_LIGHT = "#6c6c6c"
COLOR_MALE = "#5B8DEF"
COLOR_FEMALE = "#FF7E93"
COLOR_UNKNOWN = "#AAAAAA"
COLOR_PAW = "#D9A05B"
COLOR_HOVER = "#FAF7EE"
COLOR_DIVIDER = "#EEE9DD"
COLOR_DANGER = "#E57373"
COLOR_OK = "#7CB87C"
COLOR_WARN = "#E8A53A"
COLOR_INFO = "#5B8DEF"

FONT_CANDIDATES_UI = [
    "Noto Sans CJK SC", "Noto Sans SC", "Source Han Sans SC",
    "Source Han Sans CN", "思源黑体", "思源黑体 CN",
    "Sarasa Gothic SC", "Sarasa UI SC",
    "LXGW WenKai", "LXGW WenKai GB", "霞鹜文楷",
    "WenQuanYi Micro Hei", "文泉驿微米黑",
    "DejaVu Sans", "Liberation Sans",
]
FONT_CANDIDATES_MONO = [
    "JetBrains Mono", "Sarasa Mono SC", "Source Code Pro",
    "DejaVu Sans Mono", "Liberation Mono",
]

FONT_FAMILY_UI = None
FONT_FAMILY_MONO = None

FS_TITLE = 22
FS_SUBTITLE = 11
FS_CARD_NAME = 13
FS_CARD_DESC = 9
FS_BODY = 10
FS_SMALL = 9
FS_TINY = 8


# ============================================================
# 路径与字段
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "统计信息.xlsx"
JSON_PATH = BASE_DIR / "cats.json"
CLASSIFIED_DIR = BASE_DIR / "classified"
INDEX_HTML = BASE_DIR / "index.html"
TEMP_THUMB_DIR = BASE_DIR / ".tmp_thumbs"          # 临时缩略图目录

THUMB_SUFFIX = "_thumb"
THUMB_OUTPUT_SIZE = 300
THUMB_QUALITY = 90
ROW_THUMB_SIZE = 48
PREVIEW_SIZE = 320

GENDER_OPTIONS = ["male", "female", "unknown"]
STATUS_OPTIONS = ["normal", "star", "lost", "adopted"]
GENDER_LABEL = {"male": "♂ 男孩", "female": "♀ 女孩", "unknown": "❓ 未知"}
GENDER_COLOR = {"male": COLOR_MALE, "female": COLOR_FEMALE, "unknown": COLOR_UNKNOWN}
STATUS_LABEL = {"normal": "正常", "star": "⭐ 喵星", "lost": "🔍 失踪", "adopted": "🏡 被领养"}
STATUS_BADGE = {"normal": "", "star": "⭐", "lost": "🔍", "adopted": "🏡"}

COLUMN_KEYWORDS = {
    "id": "编号", "name": "姓名", "gender": "性别", "affection": "亲人指数",
    "status": "状态", "desc": "概要", "story": "故事", "pic": "图名",
}


# ============================================================
# Excel 读写
# ============================================================
class ExcelStore:
    def __init__(self, path):
        self.path = path
        self.wb = None
        self.ws = None
        self.col_idx = {}

    def load(self):
        if not self.path.exists():
            raise FileNotFoundError(f"找不到 {self.path}")
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        headers = {}
        for c in range(1, self.ws.max_column + 1):
            v = self.ws.cell(row=1, column=c).value
            if v is None:
                continue
            headers[c] = str(v).strip()
        for key, kw in COLUMN_KEYWORDS.items():
            matched = [c for c, h in headers.items() if kw in h]
            if not matched:
                raise KeyError(f"Excel 中找不到含 '{kw}' 的列")
            if len(matched) > 1:
                raise ValueError(f"列 '{kw}' 匹配到多个: {matched}")
            self.col_idx[key] = matched[0]

    def all_rows(self):
        rows = []
        for r in range(2, self.ws.max_row + 1):
            raw_id = self.ws.cell(row=r, column=self.col_idx["id"]).value
            if raw_id is None or str(raw_id).strip() == "":
                continue
            rows.append({
                "_row": r,
                "id": self._fmt_id(raw_id),
                "name": self._cell_str(r, "name"),
                "gender": self._cell_str(r, "gender") or "unknown",
                "affection": self._cell_int(r, "affection", default=1),
                "status": self._cell_str(r, "status") or "normal",
                "desc": self._cell_str(r, "desc"),
                "story": self._cell_str(r, "story"),
                "pic_name": self._cell_str(r, "pic"),
            })
        return rows

    def _cell_str(self, r, key):
        v = self.ws.cell(row=r, column=self.col_idx[key]).value
        return "" if v is None else str(v).strip()

    def _cell_int(self, r, key, default=1):
        v = self.ws.cell(row=r, column=self.col_idx[key]).value
        if v is None or str(v).strip() == "":
            return default
        try:
            return int(v)
        except Exception:
            return default

    @staticmethod
    def _fmt_id(raw):
        if isinstance(raw, (int, float)):
            return f"{int(raw):02d}"
        return str(raw).strip().zfill(2)

    def update_row(self, row_index, fields):
        for key, val in fields.items():
            if key in self.col_idx:
                self.ws.cell(row=row_index, column=self.col_idx[key]).value = val

    def find_row_by_id(self, cat_id):
        for r in range(2, self.ws.max_row + 1):
            raw = self.ws.cell(row=r, column=self.col_idx["id"]).value
            if raw is None:
                continue
            if self._fmt_id(raw) == cat_id:
                return r
        return None

    def append_row(self, fields):
        new_r = self.ws.max_row + 1
        while self.ws.cell(row=new_r, column=self.col_idx["id"]).value not in (None, ""):
            new_r += 1
        self.update_row(new_r, fields)
        return new_r

    def delete_row(self, row_index):
        self.ws.delete_rows(row_index, 1)

    def next_id(self):
        max_id = 0
        for r in range(2, self.ws.max_row + 1):
            raw = self.ws.cell(row=r, column=self.col_idx["id"]).value
            if raw is None:
                continue
            try:
                max_id = max(max_id, int(str(raw).strip()))
            except Exception:
                pass
        return f"{max_id + 1:02d}"

    def save(self):
        self.wb.save(self.path)

    def rewrite_rows(self, rows):
        while self.ws.max_row > 1:
            self.ws.delete_rows(2)
        for idx, cat in enumerate(rows, start=2):
            self.ws.cell(row=idx, column=self.col_idx["id"]).value = int(cat["id"])
            self.ws.cell(row=idx, column=self.col_idx["name"]).value = cat["name"]
            self.ws.cell(row=idx, column=self.col_idx["gender"]).value = cat["gender"]
            self.ws.cell(row=idx, column=self.col_idx["affection"]).value = cat["affection"]
            self.ws.cell(row=idx, column=self.col_idx["status"]).value = cat["status"]
            self.ws.cell(row=idx, column=self.col_idx["desc"]).value = cat["desc"]
            self.ws.cell(row=idx, column=self.col_idx["story"]).value = cat["story"]
            self.ws.cell(row=idx, column=self.col_idx["pic"]).value = cat["pic_name"]


# ============================================================
# cats.json 重生成
# ============================================================
def regenerate_cats_json(rows, classified_dir, json_path):
    warnings = []
    cats = []
    for row in rows:
        cat_id = row["id"]
        name = row["name"]
        pic_name = row["pic_name"]
        folder_name = f"{cat_id} {name}"
        folder_real = classified_dir / folder_name
        folder_json = f"classified/{folder_name}"

        if not folder_real.is_dir():
            warnings.append(f"[{cat_id}] 文件夹缺失: {folder_name}")
            continue
        if not pic_name:
            warnings.append(f"[{cat_id}] 图名为空")
            continue

        avatar = f"{folder_json}/{pic_name}_01_thumb.jpg"
        avatar_hd = f"{folder_json}/{pic_name}_01.jpg"
        if not (folder_real / f"{pic_name}_01.jpg").is_file():
            warnings.append(f"[{cat_id}] 缺头像原图 {pic_name}_01.jpg")
        if not (folder_real / f"{pic_name}_01_thumb.jpg").is_file():
            warnings.append(f"[{cat_id}] 缺头像缩略图 {pic_name}_01_thumb.jpg")

        other_photos = []
        pattern = str(folder_real / f"{pic_name}_[0-9][0-9].jpg")
        for fpath in sorted(glob.glob(pattern)):
            m = re.match(rf"^{re.escape(pic_name)}_(\d{{2}})\.jpg$", os.path.basename(fpath))
            if not m:
                continue
            seq = int(m.group(1))
            if seq < 2:
                continue
            seq_str = f"{seq:02d}"
            thumb_file = folder_real / f"{pic_name}_{seq_str}_thumb.jpg"
            if not thumb_file.is_file():
                warnings.append(f"[{cat_id}] 缺缩略图 {thumb_file.name}")
            other_photos.append({
                "thumb": f"{folder_json}/{pic_name}_{seq_str}_thumb.jpg",
                "hd": f"{folder_json}/{pic_name}_{seq_str}.jpg",
            })

        cats.append({
            "id": cat_id, "name": name,
            "gender": row["gender"] or "unknown",
            "avatar": avatar, "avatar_hd": avatar_hd,
            "affection": row["affection"],
            "status": row["status"] or "normal",
            "desc": row["desc"], "story": row["story"],
            "otherPhotos": other_photos,
        })

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(cats, f, ensure_ascii=False, indent=2)
    return len(cats), warnings


# ============================================================
# 缩略图工具
# ============================================================
def auto_thumbnail(src_path, dst_path, output_size=THUMB_OUTPUT_SIZE, crop_ratio=0.6):
    img = Image.open(src_path).convert("RGB")
    w, h = img.size
    crop = max(100, int(min(w, h) * crop_ratio))
    left = (w - crop) // 2
    top = (h - crop) // 2
    img = img.crop((left, top, left + crop, top + crop))
    img = img.resize((output_size, output_size), Image.LANCZOS)
    img.save(dst_path, "JPEG", quality=THUMB_QUALITY, optimize=True)


def next_seq_for(folder, pic_name):
    if not folder.is_dir():
        return 1
    used = set()
    for f in folder.iterdir():
        m = re.match(rf"^{re.escape(pic_name)}_(\d{{2}})(?:_thumb)?\.jpe?g$",
                     f.name, re.IGNORECASE)
        if m:
            used.add(int(m.group(1)))
    n = 1
    while n in used:
        n += 1
    return n


def open_in_explorer(path: Path):
    try:
        if sys.platform == "win32":
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception as e:
        messagebox.showerror("打不开", str(e))


def detect_fonts(root):
    global FONT_FAMILY_UI, FONT_FAMILY_MONO
    available = set(tkfont.families(root))
    for name in FONT_CANDIDATES_UI:
        if name in available:
            FONT_FAMILY_UI = name
            break
    for name in FONT_CANDIDATES_MONO:
        if name in available:
            FONT_FAMILY_MONO = name
            break

    default = tkfont.nametofont("TkDefaultFont")
    if FONT_FAMILY_UI:
        default.configure(family=FONT_FAMILY_UI, size=10)
    else:
        default.configure(size=10)
    text_font = tkfont.nametofont("TkTextFont")
    if FONT_FAMILY_UI:
        text_font.configure(family=FONT_FAMILY_UI, size=10)
    fixed = tkfont.nametofont("TkFixedFont")
    if FONT_FAMILY_MONO:
        fixed.configure(family=FONT_FAMILY_MONO, size=10)


def ui_font(size=FS_BODY, weight="normal"):
    if FONT_FAMILY_UI:
        return (FONT_FAMILY_UI, size, weight)
    return ("TkDefaultFont", size, weight)


def mono_font(size=FS_SMALL):
    if FONT_FAMILY_MONO:
        return (FONT_FAMILY_MONO, size)
    return ("TkFixedFont", size)


# ============================================================
# 缩略图裁剪器（增加 on_close 回调）
# ============================================================
class ThumbEditor(tk.Toplevel):
    def __init__(self, master, original_path, on_done=None, on_close=None):
        super().__init__(master)
        self.title(f"编辑缩略图 - {original_path.name}")
        self.geometry("1000x720")
        self.configure(bg=COLOR_BG)
        self.transient(master)
        self.grab_set()

        self.original_path = original_path
        self.on_done = on_done
        self.on_close = on_close          # 新增：关闭时回调（不论是否保存）
        self.original_image = Image.open(original_path)
        self.tk_image = None
        self.crop_box = None
        self.display_crop_box = None
        self.display_scale = 1.0
        self.img_x = 0; self.img_y = 0
        self.dragging = False; self.resizing = False
        self.resize_corner = None
        self.offset_x = 0; self.offset_y = 0
        self.crop_w = 0; self.crop_h = 0
        self.min_crop_size = 100
        self.output_size = THUMB_OUTPUT_SIZE

        self._build_ui()
        self._init_crop_box()
        self.after(80, self._display)
        self.bind("<Configure>", self._on_resize)
        self.protocol("WM_DELETE_WINDOW", self._on_window_close)

    def _on_window_close(self):
        if self.on_close:
            self.on_close()
        self.destroy()

    def _build_ui(self):
        toolbar = tk.Frame(self, bg=COLOR_ACCENT, height=56)
        toolbar.pack(fill=tk.X); toolbar.pack_propagate(False)
        tk.Label(toolbar, text=f"✏️  编辑缩略图：{self.original_path.name}",
                 font=ui_font(FS_BODY+2, "bold"), bg=COLOR_ACCENT, fg="white"
                 ).pack(side=tk.LEFT, padx=18)
        tk.Button(toolbar, text="💾 保存覆盖", command=self._save,
                  font=ui_font(FS_BODY, "bold"), bg=COLOR_WARN, fg="white",
                  padx=18, pady=6, relief=tk.FLAT, cursor="hand2", borderwidth=0
                  ).pack(side=tk.RIGHT, padx=10, pady=10)
        tk.Button(toolbar, text="↺ 重置", command=self._reset,
                  font=ui_font(FS_BODY), bg="#fff8d6", fg=COLOR_TEXT,
                  padx=14, pady=6, relief=tk.FLAT, cursor="hand2", borderwidth=0
                  ).pack(side=tk.RIGHT, padx=4, pady=10)
        tk.Button(toolbar, text="🎯 居中", command=self._center,
                  font=ui_font(FS_BODY), bg="#e3f2fd", fg=COLOR_TEXT,
                  padx=14, pady=6, relief=tk.FLAT, cursor="hand2", borderwidth=0
                  ).pack(side=tk.RIGHT, padx=4, pady=10)

        info_bar = tk.Frame(self, bg=COLOR_DIVIDER, height=32)
        info_bar.pack(fill=tk.X); info_bar.pack_propagate(False)
        self.info_label = tk.Label(info_bar, text="拖动方框移动，拖角调整大小（保持 1:1）",
                                   bg=COLOR_DIVIDER, fg=COLOR_SUBTEXT, font=ui_font(FS_SMALL))
        self.info_label.pack(side=tk.LEFT, padx=14)

        self.canvas = tk.Canvas(self, bg="#f0f0f0", highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.bind("<ButtonPress-1>", self._on_down)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_up)

    def _init_crop_box(self):
        w, h = self.original_image.size
        side = max(100, int(min(w, h) * 0.6))
        x1 = (w - side) // 2; y1 = (h - side) // 2
        self.crop_box = (x1, y1, x1 + side, y1 + side)

    def _display(self):
        cw = self.canvas.winfo_width() or 800
        ch = self.canvas.winfo_height() or 580
        iw, ih = self.original_image.size
        self.display_scale = min(cw / iw, ch / ih, 1.0)
        dw, dh = int(iw * self.display_scale), int(ih * self.display_scale)
        disp = self.original_image.resize((dw, dh), Image.LANCZOS)
        self.tk_image = ImageTk.PhotoImage(disp)
        self.canvas.delete("all")
        self.img_x = (cw - dw) // 2
        self.img_y = (ch - dh) // 2
        self.canvas.create_image(self.img_x, self.img_y, anchor=tk.NW, image=self.tk_image)
        self._draw_box()

    def _draw_box(self):
        if not self.crop_box:
            return
        x1, y1, x2, y2 = self.crop_box
        s = self.display_scale
        dx1 = self.img_x + int(x1 * s); dy1 = self.img_y + int(y1 * s)
        dx2 = self.img_x + int(x2 * s); dy2 = self.img_y + int(y2 * s)
        self.display_crop_box = (dx1, dy1, dx2, dy2)
        cw = self.canvas.winfo_width(); ch = self.canvas.winfo_height()
        self.canvas.delete("box")
        for rect in [(0, 0, dx1, ch), (dx2, 0, cw, ch),
                     (dx1, 0, dx2, dy1), (dx1, dy2, dx2, ch)]:
            self.canvas.create_rectangle(*rect, fill="#000", stipple="gray50",
                                         outline="", tags="box")
        self.canvas.create_rectangle(dx1, dy1, dx2, dy2,
                                     outline=COLOR_INFO, width=3, tags="box")
        for hx, hy in [(dx1, dy1), (dx2, dy1), (dx1, dy2), (dx2, dy2)]:
            self.canvas.create_oval(hx - 7, hy - 7, hx + 7, hy + 7,
                                    fill="#FF7043", outline="white", width=2, tags="box")
        side = x2 - x1
        self.info_label.config(text=f"裁剪 {side}×{side}  ▸  输出 {self.output_size}×{self.output_size}")

    def _on_down(self, e):
        if not self.display_crop_box: return
        dx1, dy1, dx2, dy2 = self.display_crop_box
        for name, (cx, cy) in [("nw", (dx1, dy1)), ("ne", (dx2, dy1)),
                                ("sw", (dx1, dy2)), ("se", (dx2, dy2))]:
            if abs(e.x - cx) < 12 and abs(e.y - cy) < 12:
                self.resizing = True; self.resize_corner = name
                self.offset_x, self.offset_y = e.x, e.y
                return
        if dx1 <= e.x <= dx2 and dy1 <= e.y <= dy2:
            self.dragging = True
            self.offset_x = e.x - dx1; self.offset_y = e.y - dy1
            self.crop_w = self.crop_box[2] - self.crop_box[0]
            self.crop_h = self.crop_box[3] - self.crop_box[1]

    def _on_drag(self, e):
        if self.dragging: self._move(e)
        elif self.resizing: self._resize(e)

    def _on_up(self, _):
        self.dragging = False; self.resizing = False; self.resize_corner = None

    def _move(self, e):
        s = self.display_scale
        new_dx1 = e.x - self.offset_x; new_dy1 = e.y - self.offset_y
        x1f = (new_dx1 - self.img_x) / s; y1f = (new_dy1 - self.img_y) / s
        iw, ih = self.original_image.size
        x1 = max(0, min(int(x1f), iw - self.crop_w))
        y1 = max(0, min(int(y1f), ih - self.crop_h))
        self.crop_box = (x1, y1, x1 + self.crop_w, y1 + self.crop_h)
        self._draw_box()

    def _resize(self, e):
        x1, y1, x2, y2 = self.crop_box
        iw, ih = self.original_image.size
        s = self.display_scale
        dxr = (e.x - self.offset_x) / s; dyr = (e.y - self.offset_y) / s
        if self.resize_corner == "nw": x1 += dxr; y1 += dyr
        elif self.resize_corner == "ne": x2 += dxr; y1 += dyr
        elif self.resize_corner == "sw": x1 += dxr; y2 += dyr
        elif self.resize_corner == "se": x2 += dxr; y2 += dyr
        x1 = max(0, x1); y1 = max(0, y1)
        x2 = min(iw, x2); y2 = min(ih, y2)
        if self.resize_corner == "se": fx, fy = x1, y1
        elif self.resize_corner == "sw": fx, fy = x2, y1
        elif self.resize_corner == "ne": fx, fy = x1, y2
        else: fx, fy = x2, y2
        side = max(min(x2 - x1, y2 - y1), self.min_crop_size / s)
        if self.resize_corner == "se":
            x1, y1 = fx, fy; x2, y2 = x1 + side, y1 + side
        elif self.resize_corner == "sw":
            x2, y1 = fx, fy; x1, y2 = x2 - side, y1 + side
        elif self.resize_corner == "ne":
            x1, y2 = fx, fy; x2, y1 = x1 + side, y2 - side
        else:
            x2, y2 = fx, fy; x1, y1 = x2 - side, y2 - side
        x1 = max(0, x1); y1 = max(0, y1)
        x2 = min(iw, x2); y2 = min(ih, y2)
        self.crop_box = (int(x1), int(y1), int(x2), int(y2))
        self._draw_box()
        self.offset_x, self.offset_y = e.x, e.y

    def _on_resize(self, _e):
        if self.original_image and self.tk_image:
            self._display()

    def _reset(self):
        self._init_crop_box(); self._display()

    def _center(self):
        x1, y1, x2, y2 = self.crop_box
        size = x2 - x1
        iw, ih = self.original_image.size
        nx = (iw - size) // 2; ny = (ih - size) // 2
        self.crop_box = (nx, ny, nx + size, ny + size)
        self._display()

    def _save(self):
        try:
            cropped = self.original_image.crop(self.crop_box)
            thumb = cropped.resize((self.output_size, self.output_size), Image.LANCZOS)
            stem = self.original_path.stem
            ext = self.original_path.suffix
            out_path = self.original_path.with_name(f"{stem}{THUMB_SUFFIX}{ext}")
            thumb.save(out_path, "JPEG", quality=THUMB_QUALITY, optimize=True)
            if self.on_done:
                self.on_done(out_path)
            if self.on_close:
                self.on_close()
            self.destroy()
        except Exception as e:
            messagebox.showerror("失败", f"{e}\n\n{traceback.format_exc()}")


# ============================================================
# CatEditor（修改：支持编辑已有猫咪的图名）
# ============================================================
class CatEditor(tk.Toplevel):
    def __init__(self, master, store, cat=None, on_saved=None):
        super().__init__(master)
        self.store = store
        self.cat = cat
        self.on_saved = on_saved
        self.is_new = cat is None

        title = "新增猫咪" if self.is_new else f"编辑 {cat['id']} {cat['name']}"
        self.title(title)

        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        target_w = max(1100, min(int(sw * 0.8), 1600))
        target_h = max(750, min(int(sh * 0.8), 1000))
        self.geometry(f"{target_w}x{target_h}")
        self.minsize(max(900, int(sw * 0.55)), max(700, int(sh * 0.55)))

        self.configure(bg=COLOR_BG)
        self.transient(master)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._on_close_request)

        self.new_photo_queue = []
        self._initial_snapshot = None
        self._preview_original_img = None
        self._preview_thumb_img = None
        self._last_original = None
        self._last_thumb = None

        # 确保临时目录存在
        TEMP_THUMB_DIR.mkdir(parents=True, exist_ok=True)

        self._build()
        self._load_values()
        if not self.is_new:
            self._refresh_existing_photos()
        self._initial_snapshot = self._snapshot()
        self._made_fs_changes = False

    def _snapshot(self):
        return (
            self.var_id.get(), self.var_name.get(), self.var_pic.get(),
            self.var_gender.get(), self.var_status.get(),
            int(self.var_affection.get() or 0), self.var_desc.get(),
            self.txt_story.get("1.0", tk.END),
            tuple((str(p["src"]), p["seq"], p.get("thumb_temp", None) and str(p["thumb_temp"])) for p in self.new_photo_queue),
        )

    def _is_dirty(self):
        if self._initial_snapshot is None:
            return False
        return self._snapshot() != self._initial_snapshot or self._made_fs_changes

    def _mark_dirty(self):
        self._made_fs_changes = True

    def _on_close_request(self):
        if not self._is_dirty():
            self.destroy(); return
        ans = messagebox.askyesnocancel(
            "未保存的修改",
            "你有未保存的修改。\n\n是 ＝ 保存并退出\n否 ＝ 丢弃并退出\n取消 ＝ 继续编辑"
        )
        if ans is None:
            return
        if ans:
            self._save()
        else:
            self.destroy()

    def _build(self):
        head = tk.Frame(self, bg=COLOR_ACCENT, height=58)
        head.pack(fill=tk.X); head.pack_propagate(False)
        title = "➕ 新增一只猫咪" if self.is_new else f"✏️ {self.cat['id']} {self.cat['name']}"
        tk.Label(head, text=title, font=ui_font(FS_BODY+4, "bold"),
                 bg=COLOR_ACCENT, fg="white").pack(side=tk.LEFT, padx=22, pady=14)
        tk.Label(head,
                 text="所有字段都可修改・保存后同步 Excel 与 cats.json",
                 font=ui_font(FS_SMALL), bg=COLOR_ACCENT, fg="#cccccc"
                 ).pack(side=tk.LEFT, padx=8, pady=18)

        outer = tk.Frame(self, bg=COLOR_BG)
        outer.pack(fill=tk.BOTH, expand=True, padx=20, pady=14)

        left = tk.Frame(outer, bg=COLOR_BG)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right = tk.Frame(outer, bg=COLOR_BG)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(14, 0))

        fields = self._make_card(left, "基本信息")
        fields.pack(fill=tk.X, pady=(0, 12))

        self.var_id = tk.StringVar()
        self.var_name = tk.StringVar()
        self.var_pic = tk.StringVar()
        self.var_gender = tk.StringVar(value="unknown")
        self.var_status = tk.StringVar(value="normal")
        self.var_affection = tk.IntVar(value=3)
        self.var_desc = tk.StringVar()

        body = tk.Frame(fields, bg=COLOR_CARD)
        body.pack(fill=tk.X, padx=18, pady=12)
        body.columnconfigure(1, weight=1); body.columnconfigure(3, weight=1)

        def label(parent, text, r, c):
            tk.Label(parent, text=text, bg=COLOR_CARD, fg=COLOR_SUBTEXT,
                     font=ui_font(FS_SMALL)
                     ).grid(row=r, column=c, sticky="w", padx=4, pady=6)

        def entry(parent, var, r, c, **kw):
            e = tk.Entry(parent, textvariable=var,
                         font=ui_font(FS_BODY), relief=tk.FLAT,
                         bg="#fafafa", highlightthickness=1,
                         highlightbackground=COLOR_DIVIDER,
                         highlightcolor=COLOR_INFO, **kw)
            e.grid(row=r, column=c, sticky="we", padx=4, pady=6, ipady=4)
            return e

        label(body, "编号", 0, 0)
        e_id = entry(body, self.var_id, 0, 1, width=10)
        if not self.is_new:
            e_id.config(state="disabled")

        label(body, "姓名", 0, 2)
        entry(body, self.var_name, 0, 3)

        label(body, "图名", 1, 0)
        e_pic = entry(body, self.var_pic, 1, 1)
        # ---------- 修改点 1：图名输入框始终可编辑 ----------
        # 原代码：if not self.is_new: e_pic.config(state="disabled")
        # 现在不再禁用，但添加提示文字

        label(body, "概要", 1, 2)
        entry(body, self.var_desc, 1, 3)

        label(body, "性别", 2, 0)
        cb_g = ttk.Combobox(body, textvariable=self.var_gender,
                            values=GENDER_OPTIONS, state="readonly",
                            font=ui_font(FS_BODY))
        cb_g.grid(row=2, column=1, sticky="we", padx=4, pady=6, ipady=2)

        label(body, "状态", 2, 2)
        cb_s = ttk.Combobox(body, textvariable=self.var_status,
                            values=STATUS_OPTIONS, state="readonly",
                            font=ui_font(FS_BODY))
        cb_s.grid(row=2, column=3, sticky="we", padx=4, pady=6, ipady=2)

        label(body, "亲人指数", 3, 0)
        aff = tk.Frame(body, bg=COLOR_CARD)
        aff.grid(row=3, column=1, columnspan=3, sticky="w", padx=4, pady=6)
        tk.Spinbox(aff, from_=1, to=5, textvariable=self.var_affection,
                   font=ui_font(FS_BODY), width=5, relief=tk.FLAT,
                   bg="#fafafa", highlightthickness=1,
                   highlightbackground=COLOR_DIVIDER
                   ).pack(side=tk.LEFT, ipady=2)
        self.aff_preview = tk.Label(aff, bg=COLOR_CARD, fg=COLOR_PAW,
                                    font=ui_font(FS_BODY+2))
        self.aff_preview.pack(side=tk.LEFT, padx=10)
        self.var_affection.trace_add("write", lambda *_: self._update_paw_preview())
        self._update_paw_preview()

        # 提示文字，根据新增/编辑显示不同内容
        if self.is_new:
            tk.Label(body,
                     text="提示：编号与图名保存后不可修改。图名只能包含 英文/数字/下划线。",
                     bg=COLOR_CARD, fg=COLOR_SUBTEXT, font=ui_font(FS_TINY),
                     wraplength=520, justify="left"
                     ).grid(row=4, column=0, columnspan=4, sticky="w", padx=4, pady=(8, 0))
        else:
            # ---------- 修改点 2：编辑时提示可以修改图名，但谨慎操作 ----------
            tk.Label(body,
                     text="提示：图名修改后将重命名所有相关图片文件，请谨慎操作。只能包含英文/数字/下划线。",
                     bg=COLOR_CARD, fg=COLOR_SUBTEXT, font=ui_font(FS_TINY),
                     wraplength=520, justify="left"
                     ).grid(row=4, column=0, columnspan=4, sticky="w", padx=4, pady=(8, 0))

        story_card = self._make_card(left, "故事")
        story_card.pack(fill=tk.BOTH, expand=True, pady=(0, 0))
        story_inner = tk.Frame(story_card, bg=COLOR_CARD)
        story_inner.pack(fill=tk.BOTH, expand=True, padx=18, pady=12)
        self.txt_story = tk.Text(story_inner, height=8, font=ui_font(FS_BODY),
                                  wrap="word", relief=tk.FLAT,
                                  bg="#fafafa", padx=10, pady=10,
                                  highlightthickness=1,
                                  highlightbackground=COLOR_DIVIDER,
                                  highlightcolor=COLOR_INFO)
        self.txt_story.pack(fill=tk.BOTH, expand=True)
        tk.Label(story_inner,
                 text="空行会在前端被识别为段落分隔",
                 bg=COLOR_CARD, fg=COLOR_SUBTEXT, font=ui_font(FS_TINY)
                 ).pack(anchor="e", pady=(4, 0))

        photo_card = self._make_card(right, "图片")
        photo_card.pack(fill=tk.BOTH, expand=True)
        photo_inner = tk.Frame(photo_card, bg=COLOR_CARD)
        photo_inner.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)

        preview_frame = tk.Frame(photo_inner, bg=COLOR_CARD)
        preview_frame.pack(fill=tk.X, pady=(0, 8))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.columnconfigure(1, weight=1)

        def make_preview_box(parent, column, label_text):
            box = tk.Frame(parent, bg="#fafafa",
                           highlightthickness=1,
                           highlightbackground=COLOR_DIVIDER,
                           width=240, height=240)
            box.grid(row=0, column=column, padx=2, pady=4, sticky="n")
            box.pack_propagate(False)
            lbl = tk.Label(box, bg="#fafafa", fg=COLOR_SUBTEXT,
                           text=label_text, font=ui_font(FS_SMALL))
            lbl.pack(fill=tk.BOTH, expand=True)
            return box, lbl

        self.preview_original_box, self.preview_original_label = make_preview_box(
            preview_frame, 0, "原图预览")
        self.preview_thumb_box, self.preview_thumb_label = make_preview_box(
            preview_frame, 1, "缩略图预览")

        self.preview_original_label.bind("<Button-1>", self._open_preview_external)
        self.preview_thumb_label.bind("<Button-1>",
            lambda _e: self._open_preview_external(thumb=True))

        def pill(parent, text, color, cmd):
            return tk.Button(parent, text=text, command=cmd,
                             font=ui_font(FS_SMALL, "bold"),
                             bg=color, fg="white",
                             relief=tk.FLAT, borderwidth=0,
                             padx=10, pady=5, cursor="hand2")

        if self.is_new:
            row1 = tk.Frame(photo_inner, bg=COLOR_CARD)
            row1.pack(fill=tk.X, pady=(2, 2))
            pill(row1, "📷 选主图", COLOR_OK,
                 lambda: self._start_photo_selection(seq=1)).pack(side=tk.LEFT, padx=2)
            pill(row1, "➕ 加补充", COLOR_INFO,
                 lambda: self._start_photo_selection(seq=None)).pack(side=tk.LEFT, padx=2)

            row2 = tk.Frame(photo_inner, bg=COLOR_CARD)
            row2.pack(fill=tk.X, pady=(0, 4))
            pill(row2, "✏️ 重裁缩略", COLOR_WARN,
                 self._manual_rethumb_queued).pack(side=tk.LEFT, padx=2)
            pill(row2, "🗑 移除", COLOR_DANGER,
                 self._remove_queued).pack(side=tk.LEFT, padx=2)
        else:
            row1 = tk.Frame(photo_inner, bg=COLOR_CARD)
            row1.pack(fill=tk.X, pady=(2, 2))
            pill(row1, "✏️ 重裁缩略", COLOR_WARN,
                 self._edit_existing_thumb).pack(side=tk.LEFT, padx=2)
            pill(row1, "➕ 补充图", COLOR_INFO,
                 self._add_extra_existing).pack(side=tk.LEFT, padx=2)
            pill(row1, "🗑 删除", COLOR_DANGER,
                 self._delete_existing_photo).pack(side=tk.LEFT, padx=2)
            row2 = tk.Frame(photo_inner, bg=COLOR_CARD)
            row2.pack(fill=tk.X, pady=(0, 4))
            pill(row2, "🔄 设为主图", "#7E57C2",
                 self._swap_to_main).pack(side=tk.LEFT, padx=2)
            pill(row2, "📁 打开文件夹", COLOR_ACCENT_LIGHT,
                 lambda: open_in_explorer(self._folder())).pack(side=tk.LEFT, padx=2)

        list_holder = tk.Frame(photo_inner, bg="white",
                                highlightthickness=1,
                                highlightbackground=COLOR_DIVIDER)
        list_holder.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.photo_list = tk.Listbox(list_holder, font=mono_font(FS_SMALL),
                                      activestyle="dotbox", relief=tk.FLAT,
                                      bg="white", borderwidth=0,
                                      highlightthickness=0,
                                      selectbackground=COLOR_HOVER,
                                      selectforeground=COLOR_TEXT)
        self.photo_list.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        sb = ttk.Scrollbar(list_holder, command=self.photo_list.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.photo_list.config(yscrollcommand=sb.set)
        self.photo_list.bind("<<ListboxSelect>>", lambda _e: self._on_photo_select())
        self.photo_list.bind("<Double-Button-1>", lambda _e: self._open_preview_external())

        bottom = tk.Frame(self, bg=COLOR_BG)
        bottom.pack(fill=tk.X, padx=20, pady=(0, 16))
        tk.Button(bottom, text="取消", command=self._on_close_request,
                  font=ui_font(FS_BODY), bg="#eaeaea", fg=COLOR_TEXT,
                  relief=tk.FLAT, borderwidth=0,
                  padx=22, pady=8, cursor="hand2"
                  ).pack(side=tk.RIGHT, padx=4)
        tk.Button(bottom, text="💾 保存并同步", command=self._save,
                  font=ui_font(FS_BODY, "bold"), bg=COLOR_ACCENT, fg="white",
                  relief=tk.FLAT, borderwidth=0,
                  padx=24, pady=8, cursor="hand2"
                  ).pack(side=tk.RIGHT, padx=4)

    def _make_card(self, parent, title):
        card = tk.Frame(parent, bg=COLOR_CARD,
                        highlightthickness=1,
                        highlightbackground=COLOR_CARD_BORDER)
        title_bar = tk.Frame(card, bg=COLOR_CARD)
        title_bar.pack(fill=tk.X, padx=18, pady=(12, 4))
        tk.Label(title_bar, text=title, bg=COLOR_CARD, fg=COLOR_ACCENT,
                 font=ui_font(FS_BODY+1, "bold")).pack(side=tk.LEFT)
        sep = tk.Frame(card, bg=COLOR_DIVIDER, height=1)
        sep.pack(fill=tk.X, padx=18)
        return card

    def _update_paw_preview(self):
        try:
            n = int(self.var_affection.get())
        except Exception:
            n = 0
        n = max(0, min(5, n))
        self.aff_preview.config(text="🐾" * n + "·" * (5 - n))

    def _load_values(self):
        if self.is_new:
            self.var_id.set(self.store.next_id())
            return
        c = self.cat
        self.var_id.set(c["id"])
        self.var_name.set(c["name"])
        self.var_pic.set(c["pic_name"])
        self.var_gender.set(c["gender"] or "unknown")
        self.var_status.set(c["status"] or "normal")
        self.var_affection.set(c["affection"] or 1)
        self.var_desc.set(c["desc"])
        self.txt_story.delete("1.0", tk.END)
        self.txt_story.insert("1.0", c["story"])

    def _show_preview_both(self, original_path, thumb_path=None):
        if original_path and Path(original_path).is_file():
            try:
                img = Image.open(original_path).convert("RGB")
                img.thumbnail((220, 220), Image.LANCZOS)
                self._preview_original_img = ImageTk.PhotoImage(img)
                self.preview_original_label.config(image=self._preview_original_img, text="")
            except Exception:
                self._preview_original_img = None
                self.preview_original_label.config(image="", text="(无法加载)")
        else:
            self._preview_original_img = None
            self.preview_original_label.config(image="", text="(无原图)")

        if thumb_path and Path(thumb_path).is_file():
            try:
                img = Image.open(thumb_path).convert("RGB")
                img.thumbnail((220, 220), Image.LANCZOS)
                self._preview_thumb_img = ImageTk.PhotoImage(img)
                self.preview_thumb_label.config(image=self._preview_thumb_img, text="")
            except Exception:
                self._preview_thumb_img = None
                self.preview_thumb_label.config(image="", text="(无法加载)")
        else:
            self._preview_thumb_img = None
            self.preview_thumb_label.config(image="", text="选中图片后显示")

        self._last_original = original_path
        self._last_thumb = thumb_path

    def _open_preview_external(self, _e=None, thumb=False):
        path = (self._last_thumb if thumb else self._last_original)
        if path and Path(path).is_file():
            try:
                if sys.platform == "win32":
                    os.startfile(str(path))
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", str(path)])
                else:
                    subprocess.Popen(["xdg-open", str(path)])
            except Exception as e:
                messagebox.showerror("打不开", str(e))

    def _on_photo_select(self):
        idx = self._selected_idx()
        if idx is None:
            return
        if self.is_new:
            if 0 <= idx < len(self.new_photo_queue):
                item = self.new_photo_queue[idx]
                src = item["src"]
                thumb = item.get("thumb_temp", None)
                self._show_preview_both(str(src), str(thumb) if thumb else None)
        else:
            seq = self._selected_existing_seq()
            if seq is None:
                return
            folder = self._folder()
            pic = self.cat["pic_name"]
            original = folder / f"{pic}_{seq:02d}.jpg"
            thumb = folder / f"{pic}_{seq:02d}_thumb.jpg" if seq else None
            self._show_preview_both(
                str(original) if original.is_file() else None,
                str(thumb) if (thumb and thumb.is_file()) else None
            )

    def _selected_idx(self):
        sel = self.photo_list.curselection()
        return sel[0] if sel else None

    # -------------------------------------------------------
    # 新增图片流程：选文件 → 逐个弹出裁剪窗口
    # -------------------------------------------------------
    def _start_photo_selection(self, seq):
        files = filedialog.askopenfilenames(
            title="选择图片",
            filetypes=[("图片", "*.jpg *.jpeg *.png"), ("所有", "*.*")]
        )
        if not files:
            return

        if seq == 1:
            # 移除旧的主图
            self.new_photo_queue = [p for p in self.new_photo_queue if p["seq"] != 1]

        self._pending_files = list(files)
        self._pending_seq = seq
        self._process_next_file()

    def _process_next_file(self):
        if not self._pending_files:
            self._refresh_queue_list()
            return
        file_path = self._pending_files.pop(0)
        seq = self._pending_seq

        # 创建裁剪窗口，绑定回调
        def on_done(thumb_path):
            # 将临时缩略图复制到我们的临时目录，避免原目录被污染
            tmp_name = f"tmp_{os.path.basename(file_path)}_{os.urandom(4).hex()}.jpg"
            dest = TEMP_THUMB_DIR / tmp_name
            shutil.copyfile(thumb_path, dest)
            self.new_photo_queue.append({
                "src": Path(file_path),
                "seq": seq,
                "thumb_temp": dest
            })

        def on_close():
            # 无论是否保存，继续处理下一个
            self.after(50, self._process_next_file)

        # 构建一个虚拟的 original_path，让 ThumbEditor 保存到临时位置
        # 临时路径：在 TEMP_THUMB_DIR 下
        tmp_original = TEMP_THUMB_DIR / ("orig_" + os.path.basename(file_path))
        shutil.copyfile(file_path, tmp_original)

        editor = ThumbEditor(
            self,
            original_path=tmp_original,
            on_done=lambda p: on_done(p),
            on_close=on_close
        )
        # 如果用户直接关闭窗口，不会调用 on_done，条目就不会被添加
        # 但我们仍需要记录该文件（未裁剪）
        # 解决方法：在 on_close 中也添加条目，thumb_temp 为 None
        # 修改 on_close 来添加条目
        def on_close_full():
            if not any(p["src"] == Path(file_path) for p in self.new_photo_queue):
                self.new_photo_queue.append({
                    "src": Path(file_path),
                    "seq": seq,
                    "thumb_temp": None
                })
            self.after(50, self._process_next_file)

        # 重新绑定关闭回调（上面已定义 on_close，这里覆盖）
        editor.on_close = on_close_full
        editor.protocol("WM_DELETE_WINDOW", editor._on_window_close)

    def _refresh_queue_list(self):
        self.photo_list.delete(0, tk.END)
        for p in self.new_photo_queue:
            tag = "主图" if p["seq"] == 1 else "补充"
            status = "  [已裁剪]" if p.get("thumb_temp") else ""
            self.photo_list.insert(tk.END, f"[{tag}] {p['src'].name}{status}")

    def _manual_rethumb_queued(self):
        idx = self._selected_idx()
        if idx is None:
            messagebox.showwarning("提示", "请先选中一张图片")
            return
        item = self.new_photo_queue[idx]
        # 弹出裁剪窗口重新裁剪
        tmp_original = TEMP_THUMB_DIR / ("orig_" + item["src"].name)
        shutil.copyfile(str(item["src"]), tmp_original)

        def on_done(thumb_path):
            tmp_name = f"tmp_{item['src'].name}_{os.urandom(4).hex()}.jpg"
            dest = TEMP_THUMB_DIR / tmp_name
            shutil.copyfile(thumb_path, dest)
            item["thumb_temp"] = dest
            self._refresh_queue_list()

        def on_close():
            pass

        editor = ThumbEditor(self, tmp_original, on_done=on_done, on_close=on_close)

    def _remove_queued(self):
        idx = self._selected_idx()
        if idx is None:
            return
        del self.new_photo_queue[idx]
        self._refresh_queue_list()

    # -------------------------------------------------------
    # 已有猫咪的图片管理（保持不变）
    # -------------------------------------------------------
    def _folder(self):
        return CLASSIFIED_DIR / f"{self.cat['id']} {self.cat['name']}"

    def _refresh_existing_photos(self):
        self.photo_list.delete(0, tk.END)
        folder = self._folder()
        if not folder.is_dir():
            self.photo_list.insert(tk.END, "(文件夹不存在)")
            return
        pic = self.cat["pic_name"]
        seqs = []
        for f in folder.iterdir():
            m = re.match(rf"^{re.escape(pic)}_(\d{{2}})\.jpg$", f.name, re.IGNORECASE)
            if m:
                seqs.append(int(m.group(1)))
        for seq in sorted(set(seqs)):
            tag = "主图" if seq == 1 else "补充"
            thumb_ok = (folder / f"{pic}_{seq:02d}_thumb.jpg").is_file()
            mark = "" if thumb_ok else "  [缺缩略图]"
            self.photo_list.insert(tk.END, f"[{tag}] {pic}_{seq:02d}.jpg{mark}")
        if self.photo_list.size() > 0:
            self.photo_list.selection_set(0)
            self._on_photo_select()

    def _selected_existing_seq(self):
        idx = self._selected_idx()
        if idx is None:
            return None
        line = self.photo_list.get(idx)
        m = re.search(r"_(\d{2})\.jpg", line)
        return int(m.group(1)) if m else None

    def _edit_existing_thumb(self):
        seq = self._selected_existing_seq()
        if seq is None:
            messagebox.showwarning("提示", "请先选中一张原图")
            return
        folder = self._folder()
        original = folder / f"{self.cat['pic_name']}_{seq:02d}.jpg"
        if not original.is_file():
            messagebox.showerror("失败", f"原图不存在: {original.name}")
            return
        ThumbEditor(self, original, on_done=lambda _p: (self._refresh_existing_photos(), self._mark_dirty()))

    def _add_extra_existing(self):
        files = filedialog.askopenfilenames(
            title="选择补充原图",
            filetypes=[("图片", "*.jpg *.jpeg *.png"), ("所有", "*.*")]
        )
        if not files:
            return
        folder = self._folder()
        folder.mkdir(parents=True, exist_ok=True)
        pic = self.cat["pic_name"]
        added = []
        for src in files:
            seq = next_seq_for(folder, pic)
            if seq < 2:
                seq = 2
                while (folder / f"{pic}_{seq:02d}.jpg").exists():
                    seq += 1
            dst = folder / f"{pic}_{seq:02d}.jpg"
            try:
                Image.open(src).convert("RGB").save(dst, "JPEG", quality=95)
                thumb_dst = folder / f"{pic}_{seq:02d}_thumb.jpg"
                auto_thumbnail(dst, thumb_dst)
                added.append(dst.name)
            except Exception as e:
                messagebox.showerror("复制失败", f"{src}: {e}")
        self._refresh_existing_photos()
        if added:
            messagebox.showinfo("已添加", "复制并生成缩略图：\n" + "\n".join(added))
        self._mark_dirty()

    def _delete_existing_photo(self):
        seq = self._selected_existing_seq()
        if seq is None:
            messagebox.showwarning("提示", "请先选中一张")
            return
        if seq == 1:
            messagebox.showwarning("拒绝", "主图 (_01) 不能在这里删除")
            return
        if not messagebox.askyesno("确认", f"删除 _{seq:02d} 原图与缩略图？"):
            return
        folder = self._folder()
        pic = self.cat["pic_name"]
        targets = [folder / f"{pic}_{seq:02d}.jpg",
                   folder / f"{pic}_{seq:02d}_thumb.jpg"]
        deleted = []
        for t in targets:
            if t.is_file():
                try:
                    t.unlink(); deleted.append(t.name)
                except Exception as e:
                    messagebox.showerror("删除失败", f"{t}: {e}")
        self._refresh_existing_photos()
        messagebox.showinfo("已删除", "\n".join(deleted) or "无变化")
        self._mark_dirty()

    def _swap_to_main(self):
        seq = self._selected_existing_seq()
        if seq is None:
            messagebox.showwarning("提示", "请先选中一张补充照片")
            return
        if seq == 1:
            messagebox.showwarning("提示", "这已经是主图了")
            return
        folder = self._folder()
        pic = self.cat["pic_name"]
        if not messagebox.askyesno("确认切换",
            f"将 {pic}_{seq:02d} 设为主图，\n并与原主图 {pic}_01 交换文件名？"):
            return
        renamed = []
        def swap_pair(a_seq, b_seq):
            for suffix in [".jpg", "_thumb.jpg"]:
                a_file = folder / f"{pic}_{a_seq:02d}{suffix}"
                b_file = folder / f"{pic}_{b_seq:02d}{suffix}"
                tmp_file = folder / f"{pic}_{a_seq:02d}{suffix}.tmpswap"
                if a_file.is_file():
                    a_file.rename(tmp_file)
                if b_file.is_file():
                    b_file.rename(a_file)
                if tmp_file.is_file():
                    tmp_file.rename(b_file)
        try:
            swap_pair(1, seq)
            self._refresh_existing_photos()
            self._on_photo_select()
            messagebox.showinfo("已切换",
                f"主图已切换：{pic}_{seq:02d} ↔ {pic}_01\n\n"
                "前端需点击「应用变更」刷新。")
        except Exception as e:
            messagebox.showerror("切换失败", str(e))
            self._refresh_existing_photos()
        self._mark_dirty()

    # ---------- 修改点 3：_save 中支持图名变更 ----------
    def _save(self):
        try:
            cat_id = self.var_id.get().strip().zfill(2)
            name = self.var_name.get().strip()
            pic = self.var_pic.get().strip()
            gender = (self.var_gender.get().strip() or "unknown")
            status = (self.var_status.get().strip() or "normal")
            try:
                affection = int(self.var_affection.get())
            except Exception:
                affection = 1
            desc = self.var_desc.get().strip()
            story = self.txt_story.get("1.0", tk.END).strip()

            if not name:
                messagebox.showwarning("提示", "姓名不能为空")
                return
            if not pic:
                messagebox.showwarning("提示", "图名不能为空")
                return
            if not re.match(r"^[A-Za-z0-9_]+$", pic):
                messagebox.showwarning("提示", "图名只能含 英文/数字/下划线")
                return

            actions = []
            folder = CLASSIFIED_DIR / f"{cat_id} {name}"

            if self.is_new:
                # 新增流程保持不变
                if folder.exists():
                    messagebox.showerror("失败", f"文件夹已存在：{folder.name}")
                    return
                if not any(p["seq"] == 1 for p in self.new_photo_queue):
                    messagebox.showwarning("提示", "请先选主图")
                    return
                folder.mkdir(parents=True, exist_ok=True)
                actions.append(f"创建文件夹: {folder.name}")

                main_item = next(p for p in self.new_photo_queue if p["seq"] == 1)
                others = [p for p in self.new_photo_queue if p["seq"] != 1]
                self._copy_one(main_item, folder, pic, 1, actions)
                seq = 2
                for it in others:
                    while (folder / f"{pic}_{seq:02d}.jpg").exists():
                        seq += 1
                    self._copy_one(it, folder, pic, seq, actions)
                    seq += 1

                self.store.append_row({
                    "id": int(cat_id),
                    "name": name, "gender": gender, "affection": affection,
                    "status": status, "desc": desc, "story": story, "pic": pic,
                })
                actions.append(f"追加 Excel 行 编号={cat_id}")
            else:
                # ---------- 编辑已有猫咪 ----------
                # 1. 如果图名发生变化，先在旧文件夹内重命名所有图片文件
                old_pic = self.cat["pic_name"]
                if pic != old_pic:
                    old_folder = CLASSIFIED_DIR / f"{self.cat['id']} {self.cat['name']}"
                    if not old_folder.is_dir():
                        messagebox.showerror("失败", f"猫咪文件夹不存在: {old_folder.name}")
                        return

                    # 收集需要重命名的文件
                    rename_ops = []  # (旧路径, 新路径)
                    try:
                        for f in old_folder.iterdir():
                            m = re.match(rf"^{re.escape(old_pic)}_(\d{{2}})(?:_thumb)?\.jpg$",
                                         f.name, re.IGNORECASE)
                            if m:
                                seq = m.group(1)
                                is_thumb = "_thumb" in f.name
                                new_name = f"{pic}_{seq}{'_thumb' if is_thumb else ''}.jpg"
                                new_path = old_folder / new_name
                                rename_ops.append((f, new_path))

                        # 执行重命名
                        for old_path, new_path in rename_ops:
                            old_path.rename(new_path)

                        actions.append(f"图名重命名: {old_pic} → {pic} ({len(rename_ops)} 个文件)")
                    except Exception as e:
                        # 回滚已经重命名的文件
                        for (old_path, new_path) in reversed(rename_ops):
                            if new_path.exists():
                                try:
                                    new_path.rename(old_path)
                                except Exception:
                                    pass
                        messagebox.showerror("图名修改失败",
                                             f"重命名图片时出错，已回滚所有更改：{e}")
                        return

                # 2. 如果姓名发生变化，重命名文件夹（旧逻辑）
                old_folder = self._folder()
                if name != self.cat["name"] and old_folder.is_dir():
                    if folder.exists():
                        messagebox.showerror("失败", f"目标文件夹已存在：{folder.name}")
                        return
                    old_folder.rename(folder)
                    actions.append(f"重命名文件夹: {old_folder.name} → {folder.name}")

                # 3. 更新 Excel
                row_idx = self.store.find_row_by_id(cat_id)
                if row_idx is None:
                    messagebox.showerror("失败", f"Excel 中找不到编号 {cat_id}")
                    return
                self.store.update_row(row_idx, {
                    "name": name, "gender": gender, "affection": affection,
                    "status": status, "desc": desc, "story": story, "pic": pic,
                })
                actions.append(f"更新 Excel 行 编号={cat_id}")

            # 清理临时缩略图文件
            for item in self.new_photo_queue:
                thumb_temp = item.get("thumb_temp")
                if thumb_temp and thumb_temp.exists():
                    try:
                        thumb_temp.unlink()
                    except Exception:
                        pass
            self.result_actions = actions
            self._initial_snapshot = self._snapshot()
            if self.on_saved:
                self.on_saved(actions)
            self.destroy()
        except Exception as e:
            messagebox.showerror("保存失败", f"{e}\n\n{traceback.format_exc()}")

    def _copy_one(self, item, folder, pic, seq, actions):
        src = item["src"]
        dst = folder / f"{pic}_{seq:02d}.jpg"
        Image.open(src).convert("RGB").save(dst, "JPEG", quality=95)
        actions.append(f"复制原图 → {dst.name}")

        thumb_dst = folder / f"{pic}_{seq:02d}_thumb.jpg"
        if item.get("thumb_temp") and item["thumb_temp"].exists():
            # 使用手动裁剪的结果
            shutil.copyfile(str(item["thumb_temp"]), thumb_dst)
            actions.append(f"使用手动裁剪缩略图 → {thumb_dst.name}")
        else:
            auto_thumbnail(dst, thumb_dst)
            actions.append(f"自动生成缩略图 → {thumb_dst.name}")


# ============================================================
# 主窗口 —— 行列表 + 上下移步数 + 内建预览服务器 + 关闭提醒
# ============================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🐱 SUAT-cats 管理器")
        self.geometry("1280x820")
        self.minsize(960, 640)
        self.configure(bg=COLOR_BG)

        detect_fonts(self)
        self._configure_ttk_styles()

        self.store = ExcelStore(EXCEL_PATH)
        self.store.load()
        self.rows = []
        self.thumb_cache = {}
        self.current_filter = "all"
        self.search_term = ""

        # 顺序变更脏标记
        self._order_dirty = False

        # 内建预览服务器
        self.http_server = None
        self.http_server_thread = None

        self._build_menu()
        self._build()
        self._reload_from_excel()

        # 退出时关闭服务器
        self.protocol("WM_DELETE_WINDOW", self._on_close_app)

    def _on_close_app(self):
        # 如果有未保存的顺序变更，提醒用户
        if self._order_dirty:
            ans = messagebox.askyesnocancel(
                "未保存的顺序变更",
                "你移动过猫咪的顺序，但尚未点击「应用变更」保存到 Excel。\n\n"
                "是 ＝ 保存并退出\n否 ＝ 丢弃变更并退出\n取消 ＝ 继续编辑"
            )
            if ans is None:  # 取消
                return
            if ans:  # 是 —— 保存
                try:
                    self._apply_changes_silent()
                except Exception as e:
                    messagebox.showerror("保存失败", f"自动保存顺序时出错：{e}\n\n{traceback.format_exc()}")
                    return
            # 如果 ans 为 False (否) 直接退出，不保存
        if self.http_server:
            self._stop_server()
        self.destroy()

    def _apply_changes_silent(self):
        """内部保存方法，用于关闭时自动保存，与 _apply_changes 逻辑相同但不弹窗"""
        if not self._order_dirty:
            return
        # 重新编号并重命名文件夹（与 _apply_changes 共享实现）
        self._do_apply_changes(silent=True)

    def _do_apply_changes(self, silent=False):
        """执行顺序重编号与文件夹重命名的核心方法"""
        if not self.rows:
            return

        # 1. 生成新旧 ID 映射
        old_to_new = {}
        new_to_old = {}
        for idx, cat in enumerate(self.rows, start=1):
            new_id = f"{idx:02d}"
            old_id = cat["id"]
            old_to_new[old_id] = new_id
            new_to_old[new_id] = old_id

        # 2. 检查文件夹冲突（新文件夹名是否已被其他猫占用）
        conflicts = []
        for idx, cat in enumerate(self.rows):
            new_id = f"{idx+1:02d}"
            new_folder_name = f"{new_id} {cat['name']}"
            new_folder = CLASSIFIED_DIR / new_folder_name
            old_folder = CLASSIFIED_DIR / f"{cat['id']} {cat['name']}"
            if new_folder.exists() and new_folder != old_folder:
                conflicts.append(f"{new_folder_name} (已被 {cat['id']} {cat['name']} 占用)")
        if conflicts:
            messagebox.showerror("重命名冲突",
                                 "以下文件夹名称与目标冲突，请手动处理后再试：\n" + "\n".join(conflicts))
            return

        # 3. 确认执行（仅非静默模式）
        if not silent:
            if not messagebox.askyesno("确认重新编号",
                                       f"将按当前顺序给 {len(self.rows)} 只猫咪重新编号 (01~{len(self.rows):02d})，\n"
                                       "同时重命名对应的文件夹。\n\n是否继续？"):
                return

        # 4. 先重命名所有文件夹到临时名称（避免冲突），记录操作
        rename_ops = []  # [(old_path, new_path)]
        temp_suffix = "_renaming"
        for cat in self.rows:
            old_folder = CLASSIFIED_DIR / f"{cat['id']} {cat['name']}"
            if not old_folder.is_dir():
                continue
            temp_folder = CLASSIFIED_DIR / f"{cat['id']} {cat['name']}{temp_suffix}"
            try:
                old_folder.rename(temp_folder)
                rename_ops.append((temp_folder, None))  # 待更新目标
            except Exception as e:
                # 回滚已重命名的
                for t, _ in reversed(rename_ops):
                    try:
                        t.rename(CLASSIFIED_DIR / t.name.replace(temp_suffix, ""))
                    except Exception:
                        pass
                messagebox.showerror("重命名失败", f"失败于 {old_folder.name}: {e}")
                return

        # 5. 从临时名称改为目标名称
        final_errors = []
        for idx, cat in enumerate(self.rows):
            new_id = f"{idx+1:02d}"
            temp_folder = CLASSIFIED_DIR / f"{cat['id']} {cat['name']}{temp_suffix}"
            new_folder = CLASSIFIED_DIR / f"{new_id} {cat['name']}"
            if temp_folder.is_dir():
                try:
                    temp_folder.rename(new_folder)
                except Exception as e:
                    final_errors.append(f"{temp_folder.name} → {new_folder.name}: {e}")

        if final_errors:
            messagebox.showerror("部分重命名失败",
                                 "以下文件夹未能成功重命名：\n" + "\n".join(final_errors))
            return

        # 6. 更新内存中的 id
        for idx, cat in enumerate(self.rows):
            cat["id"] = f"{idx+1:02d}"

        # 7. 保存到 Excel
        self.store.rewrite_rows(self.rows)
        self.store.save()

        # 8. 重新生成 cats.json
        count, warns = regenerate_cats_json(self.rows, CLASSIFIED_DIR, JSON_PATH)

        self._order_dirty = False

        if not silent:
            msg = f"重新编号成功，共 {count} 只猫咪。"
            if warns:
                msg += "\n\n⚠️ 警告：\n" + "\n".join(warns[:10])
                if len(warns) > 10:
                    msg += f"\n  …以及其他 {len(warns) - 10} 条"
            messagebox.showinfo("应用完成", msg)
            self.status_var.set("顺序已应用，编号已同步")
        else:
            self.status_var.set("已自动保存顺序")

        # 刷新界面
        self._reload_from_excel()

    def _apply_changes(self):
        self._do_apply_changes(silent=False)

    def _configure_ttk_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TCombobox",
                        font=ui_font(FS_BODY),
                        padding=4)
        style.configure("Vertical.TScrollbar", background=COLOR_BG,
                        troughcolor=COLOR_BG, arrowcolor=COLOR_SUBTEXT)

    def _build_menu(self):
        bar = tk.Menu(self)
        m_file = tk.Menu(bar, tearoff=False)
        m_file.add_command(label="刷新数据", command=self._reload_from_excel)
        m_file.add_command(label="手动同步前端 (cats.json)",
                           command=self._apply_changes)
        m_file.add_separator()
        m_file.add_command(label="打开 Excel 表格",
                           command=lambda: open_in_explorer(EXCEL_PATH))
        m_file.add_command(label="打开 cats.json",
                           command=lambda: open_in_explorer(JSON_PATH))
        m_file.add_command(label="打开 classified 文件夹",
                           command=lambda: open_in_explorer(CLASSIFIED_DIR))
        m_file.add_separator()
        m_file.add_command(label="退出", command=self._on_close_app)
        bar.add_cascade(label="文件", menu=m_file)

        m_tools = tk.Menu(bar, tearoff=False)
        m_tools.add_command(label="重生成所有缺失的缩略图",
                            command=self._batch_regen_missing_thumbs)
        m_tools.add_command(label="重生成全部缩略图（谨慎）",
                            command=self._batch_regen_all_thumbs)
        m_tools.add_separator()
        m_tools.add_command(label="检测表格与文件一致性",
                            command=self._check_consistency)
        bar.add_cascade(label="工具", menu=m_tools)

        m_help = tk.Menu(bar, tearoff=False)
        m_help.add_command(label="关于", command=self._show_about)
        bar.add_cascade(label="帮助", menu=m_help)

        self.config(menu=bar)

    def _show_about(self):
        info = (
            "SUAT-cats 管理器\n\n"
            "一站式管理你的猫咪档案。\n"
            "数据源：统计信息.xlsx → cats.json → index.html\n\n"
            "依赖：Pillow、openpyxl\n"
            "字体：仅使用开源字体\n\n"
            f"当前使用字体：{FONT_FAMILY_UI or 'Tk 默认'}"
        )
        messagebox.showinfo("关于", info)

    # ---------- 主体 UI ----------
    def _build(self):
        head = tk.Frame(self, bg=COLOR_BG)
        head.pack(fill=tk.X, pady=(20, 4))
        tk.Label(head, text="🐾 SUAT 🐱",
                 font=ui_font(FS_TITLE, "bold"),
                 bg=COLOR_BG, fg=COLOR_ACCENT).pack()
        tk.Label(head, text="记录每一个相遇与告别",
                 font=ui_font(FS_SUBTITLE),
                 bg=COLOR_BG, fg=COLOR_SUBTEXT).pack(pady=(2, 10))

        # 工具栏
        bar = tk.Frame(self, bg=COLOR_BG)
        bar.pack(fill=tk.X, padx=24, pady=(0, 6))

        def primary(parent, text, color, cmd):
            return tk.Button(parent, text=text, command=cmd,
                             font=ui_font(FS_BODY, "bold"),
                             bg=color, fg="white",
                             relief=tk.FLAT, borderwidth=0,
                             padx=14, pady=7, cursor="hand2")

        primary(bar, "➕ 新增猫咪", COLOR_ACCENT,
                self._add_cat).pack(side=tk.LEFT, padx=(0, 6))
        primary(bar, "🔄 刷新", COLOR_INFO,
                self._reload_from_excel).pack(side=tk.LEFT, padx=4)
        primary(bar, "💾 应用变更", COLOR_WARN,
                self._apply_changes).pack(side=tk.LEFT, padx=4)

        # 预览服务器按钮
        self.preview_btn = tk.Button(bar, text="🖥️ 预览前端",
                                     font=ui_font(FS_BODY, "bold"),
                                     bg=COLOR_OK, fg="white",
                                     relief=tk.FLAT, borderwidth=0,
                                     padx=14, pady=7, cursor="hand2",
                                     command=self._toggle_server)
        self.preview_btn.pack(side=tk.LEFT, padx=4)

        # 搜索框
        right = tk.Frame(bar, bg=COLOR_BG)
        right.pack(side=tk.RIGHT)
        self.var_search = tk.StringVar()
        e = tk.Entry(right, textvariable=self.var_search,
                     font=ui_font(FS_BODY), width=22, relief=tk.FLAT,
                     bg="white", highlightthickness=1,
                     highlightbackground=COLOR_DIVIDER,
                     highlightcolor=COLOR_INFO)
        e.pack(side=tk.RIGHT, ipady=5, padx=(4, 0))
        e.bind("<KeyRelease>", lambda _e: self._on_search())
        tk.Label(right, text="🔍", bg=COLOR_BG, fg=COLOR_SUBTEXT,
                 font=ui_font(FS_BODY)).pack(side=tk.RIGHT)

        # 筛选胶囊
        filter_bar = tk.Frame(self, bg=COLOR_BG)
        filter_bar.pack(fill=tk.X, padx=24, pady=(2, 6))
        self.filter_buttons = {}
        for label, key in [("全部猫咪", "all"),
                            ("♂ 男孩", "male"),
                            ("♀ 女孩", "female"),
                            ("❓ 未知", "unknown"),
                            ("⭐ 喵星", "star"),
                            ("🔍 失踪", "lost"),
                            ("🏡 被领养", "adopted")]:
            b = tk.Button(filter_bar, text=label,
                          font=ui_font(FS_SMALL, "bold"),
                          bg="white", fg=COLOR_SUBTEXT,
                          relief=tk.FLAT, borderwidth=0,
                          padx=14, pady=6, cursor="hand2",
                          command=lambda k=key: self._set_filter(k))
            b.pack(side=tk.LEFT, padx=3)
            self.filter_buttons[key] = b
        self._update_filter_buttons()

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status = tk.Label(self, textvariable=self.status_var,
                          bg=COLOR_DIVIDER, fg=COLOR_ACCENT_LIGHT,
                          font=ui_font(FS_SMALL),
                          anchor="w", padx=18, pady=4)
        status.pack(fill=tk.X, side=tk.BOTTOM)

        # 列表区
        outer = tk.Frame(self, bg=COLOR_BG)
        outer.pack(fill=tk.BOTH, expand=True, padx=20, pady=(8, 14))
        self.canvas = tk.Canvas(outer, bg=COLOR_BG, highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(outer, orient="vertical", command=self.canvas.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=sb.set)
        self.list_frame = tk.Frame(self.canvas, bg=COLOR_BG)
        self._list_window_id = self.canvas.create_window((0, 0), window=self.list_frame, anchor="nw")
        self.list_frame.bind("<Configure>",
                             lambda _e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>",
                          lambda e: self.canvas.itemconfig(self._list_window_id, width=e.width))
        self.canvas.bind_all("<MouseWheel>",
                             lambda e: self.canvas.yview_scroll(int(-e.delta / 120), "units"))

    # ---------- 交互 ----------
    def _set_filter(self, key):
        self.current_filter = key
        self._update_filter_buttons()
        self._render()

    def _update_filter_buttons(self):
        for k, b in self.filter_buttons.items():
            if k == self.current_filter:
                b.config(bg=COLOR_ACCENT, fg="white")
            else:
                b.config(bg="white", fg=COLOR_SUBTEXT)

    def _on_search(self):
        self.search_term = self.var_search.get().strip().lower()
        self._render()

    def _add_cat(self):
        editor = CatEditor(self, self.store, cat=None,
                            on_saved=self._on_editor_saved)
        self.wait_window(editor)

    def _edit_cat(self, cat):
        editor = CatEditor(self, self.store, cat=cat,
                            on_saved=self._on_editor_saved)
        self.wait_window(editor)

    def _on_editor_saved(self, actions):
        try:
            self.store.save()
            actions.append(f"保存 Excel: {EXCEL_PATH.name}")
            self._reload_from_excel()
            count, warns = regenerate_cats_json(self.rows, CLASSIFIED_DIR, JSON_PATH)
            actions.append(f"重生成 cats.json，共 {count} 只猫")
            msg = "\n".join("• " + a for a in actions)
            if warns:
                msg += "\n\n⚠️ 警告：\n" + "\n".join("  - " + w for w in warns[:10])
                if len(warns) > 10:
                    msg += f"\n  …以及其他 {len(warns) - 10} 条"
            messagebox.showinfo("应用成功", msg)
            self.status_var.set("已同步 Excel 与 cats.json")
            self._order_dirty = False
        except Exception as e:
            messagebox.showerror("同步失败", f"{e}\n\n{traceback.format_exc()}")

    def _reload_from_excel(self):
        try:
            self.store.load()
            self.rows = self.store.all_rows()
            self.thumb_cache.clear()
            self._render()
            self.status_var.set(f"已加载 {len(self.rows)} 只猫咪")
            self._order_dirty = False
        except Exception as e:
            messagebox.showerror("加载失败", f"{e}\n\n{traceback.format_exc()}")

    # ---------- 删除猫咪 ----------
    def _delete_cat(self, cat):
        """删除一只猫咪：确认后删除 Excel 行、整个文件夹和所有照片，然后刷新并同步 JSON。"""
        cat_id = cat['id']
        cat_name = cat['name']
        folder_path = CLASSIFIED_DIR / f"{cat_id} {cat_name}"
        # 确认对话框
        if not messagebox.askyesno(
            "⚠️ 永久删除",
            f"确定要删除 {cat_id} {cat_name} 吗？\n\n"
            f"此操作将永久删除：\n"
            f"  • Excel 中的该行数据\n"
            f"  • 整个文件夹及其所有图片：\n    {folder_path}\n\n"
            "此操作不可撤销！"
        ):
            return

        errors = []
        # 1. 删除文件夹
        if folder_path.exists():
            try:
                shutil.rmtree(folder_path)
            except Exception as e:
                errors.append(f"删除文件夹失败: {e}")

        # 2. 删除 Excel 行
        try:
            row_index = cat.get("_row")
            if row_index:
                self.store.delete_row(row_index)
                self.store.save()
            else:
                errors.append("找不到该猫咪在 Excel 中的行号")
        except Exception as e:
            errors.append(f"删除 Excel 行失败: {e}")

        if errors:
            messagebox.showerror("删除失败", "部分操作未完成：\n" + "\n".join(errors))
            self._reload_from_excel()  # 尝试恢复界面
            return

        # 3. 重新加载数据并更新前端文件
        self._reload_from_excel()
        try:
            count, warns = regenerate_cats_json(self.rows, CLASSIFIED_DIR, JSON_PATH)
            status_msg = f"已删除 {cat_id} {cat_name}，同步完成。"
            if warns:
                status_msg += f" (警告 {len(warns)} 条)"
            self.status_var.set(status_msg)
        except Exception as e:
            self.status_var.set(f"已删除 {cat_id} {cat_name}，但 cats.json 同步失败：{e}")

        # 清除顺序脏标记（删除操作已经直接保存，不再有未应用顺序）
        self._order_dirty = False

    # ---------- 过滤 ----------
    def _filtered(self):
        rs = list(self.rows)
        f = self.current_filter
        if f in GENDER_OPTIONS:
            rs = [r for r in rs if r["gender"] == f]
        elif f in ("star", "lost", "adopted"):
            rs = [r for r in rs if r["status"] == f]
        if self.search_term:
            t = self.search_term
            rs = [r for r in rs
                   if t in r["name"].lower()
                   or t in r["pic_name"].lower()
                   or t in r["desc"].lower()
                   or t in r["id"]]
        return rs

    # ---------- 缩略图加载 ----------
    def _load_small_thumb(self, cat):
        path = (CLASSIFIED_DIR / f"{cat['id']} {cat['name']}"
                / f"{cat['pic_name']}_01_thumb.jpg")
        if path in self.thumb_cache:
            return self.thumb_cache[path]
        try:
            if path.is_file():
                img = Image.open(path).convert("RGB")
            else:
                img = Image.new("RGB", (ROW_THUMB_SIZE, ROW_THUMB_SIZE), "#dddddd")
            img.thumbnail((ROW_THUMB_SIZE, ROW_THUMB_SIZE), Image.LANCZOS)
            ph = ImageTk.PhotoImage(img)
        except Exception:
            img = Image.new("RGB", (ROW_THUMB_SIZE, ROW_THUMB_SIZE), "#dddddd")
            ph = ImageTk.PhotoImage(img)
        self.thumb_cache[path] = ph
        return ph

    # ---------- 渲染行列表 ----------
    def _render(self):
        for w in self.list_frame.winfo_children():
            w.destroy()
        items = self._filtered()
        if not items:
            tk.Label(self.list_frame, text="没有符合条件的猫咪",
                     bg=COLOR_BG, fg=COLOR_SUBTEXT,
                     font=ui_font(FS_BODY+1), pady=60).pack()
            return

        for i, cat in enumerate(items):
            row_bg = COLOR_CARD if i % 2 == 0 else "#fcfcf9"
            row_frame = tk.Frame(self.list_frame, bg=row_bg, height=64,
                                 highlightthickness=0, relief=tk.FLAT)
            row_frame.pack(fill=tk.X, ipady=4, pady=1)

            # 头像缩略图
            thumb = self._load_small_thumb(cat)
            avatar = tk.Label(row_frame, image=thumb, bg=row_bg,
                              borderwidth=0, cursor="hand2")
            avatar.image = thumb
            avatar.pack(side=tk.LEFT, padx=(8, 12))
            avatar.bind("<Double-Button-1>", lambda _e, c=cat: self._edit_cat(c))

            # 信息区
            info_frame = tk.Frame(row_frame, bg=row_bg)
            info_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

            # 第一行：编号 + 姓名 + 性别 + 状态
            line1 = tk.Frame(info_frame, bg=row_bg)
            line1.pack(fill=tk.X, anchor="w")
            tk.Label(line1, text=f"{cat['id']} {cat['name']}",
                     bg=row_bg, fg=COLOR_TEXT,
                     font=ui_font(FS_BODY, "bold")).pack(side=tk.LEFT)
            gender_char = {"male": "♂", "female": "♀", "unknown": "?"}.get(cat["gender"], "?")
            gender_color = GENDER_COLOR.get(cat["gender"], COLOR_UNKNOWN)
            tk.Label(line1, text=" " + gender_char,
                     bg=row_bg, fg=gender_color,
                     font=ui_font(FS_BODY, "bold")).pack(side=tk.LEFT)
            status_label = STATUS_LABEL.get(cat["status"], "")
            if status_label:
                tk.Label(line1, text=" " + status_label,
                         bg=row_bg, fg=COLOR_SUBTEXT,
                         font=ui_font(FS_SMALL)).pack(side=tk.LEFT, padx=4)

            # 第二行：亲人指数 + 概要
            line2 = tk.Frame(info_frame, bg=row_bg)
            line2.pack(fill=tk.X, anchor="w")
            n = cat["affection"]
            paws = "🐾" * n + "·" * (5 - n)
            tk.Label(line2, text=paws, bg=row_bg, fg=COLOR_PAW,
                     font=ui_font(FS_SMALL)).pack(side=tk.LEFT)
            desc = (cat["desc"] or "···")[:30]
            tk.Label(line2, text=" " + desc, bg=row_bg, fg=COLOR_SUBTEXT,
                     font=ui_font(FS_SMALL)).pack(side=tk.LEFT)

            # 操作按钮区
            btn_frame = tk.Frame(row_frame, bg=row_bg)
            btn_frame.pack(side=tk.RIGHT, padx=4)

            # 移动步数输入框
            move_var = tk.StringVar(value="1")
            move_entry = tk.Entry(btn_frame, textvariable=move_var,
                                  width=4, font=ui_font(FS_SMALL),
                                  relief=tk.FLAT, bg="#f0f0f0",
                                  justify="center",
                                  highlightthickness=1,
                                  highlightbackground=COLOR_DIVIDER,
                                  highlightcolor=COLOR_INFO)
            move_entry.pack(side=tk.LEFT, padx=(2, 0))

            # ▲ 上移按钮
            up_btn = tk.Button(btn_frame, text="▲",
                               font=ui_font(FS_SMALL, "bold"),
                               bg=COLOR_HOVER, fg=COLOR_ACCENT,
                               relief=tk.FLAT, borderwidth=0, padx=6,
                               cursor="hand2",
                               command=lambda c=cat, v=move_var, idx=i: self._move_row(idx, v.get(), "up"))
            up_btn.pack(side=tk.LEFT, padx=1)

            # ▼ 下移按钮
            down_btn = tk.Button(btn_frame, text="▼",
                                 font=ui_font(FS_SMALL, "bold"),
                                 bg=COLOR_HOVER, fg=COLOR_ACCENT,
                                 relief=tk.FLAT, borderwidth=0, padx=6,
                                 cursor="hand2",
                                 command=lambda c=cat, v=move_var, idx=i: self._move_row(idx, v.get(), "down"))
            down_btn.pack(side=tk.LEFT, padx=1)

            # 编辑按钮
            edit_btn = tk.Button(btn_frame, text="✏️ 编辑",
                                 font=ui_font(FS_SMALL, "bold"),
                                 bg=COLOR_HOVER, fg=COLOR_ACCENT,
                                 relief=tk.FLAT, borderwidth=0, padx=8,
                                 cursor="hand2",
                                 command=lambda c=cat: self._edit_cat(c))
            edit_btn.pack(side=tk.LEFT, padx=(4, 0))

            # 删除按钮（新增）
            delete_btn = tk.Button(btn_frame, text="🗑️ 删除",
                                   font=ui_font(FS_SMALL, "bold"),
                                   bg=COLOR_DANGER, fg="white",
                                   relief=tk.FLAT, borderwidth=0, padx=8,
                                   cursor="hand2",
                                   command=lambda c=cat: self._delete_cat(c))
            delete_btn.pack(side=tk.LEFT, padx=(4, 0))

            # 分隔线
            tk.Frame(self.list_frame, bg=COLOR_DIVIDER, height=1).pack(fill=tk.X)

    # ---------- 移动行逻辑（带滚动至移动后位置）----------
    def _move_row(self, index, step_str, direction):
        try:
            step = int(step_str)
        except ValueError:
            messagebox.showwarning("输入错误", "请输入一个整数（如 1、2、-1）")
            return

        if step < 0:
            step = -step
            direction = "down" if direction == "up" else "up"

        filtered = self._filtered()
        total = len(filtered)
        if total <= 1:
            return

        cat = filtered[index]
        real_index = self.rows.index(cat)

        if direction == "up":
            new_index = max(0, real_index - step)
        else:
            new_index = min(len(self.rows) - 1, real_index + step)

        if new_index == real_index:
            return

        # 执行移动
        item = self.rows.pop(real_index)
        self.rows.insert(new_index, item)

        self.status_var.set(
            f"已将 {cat['id']} {cat['name']} 从第 {real_index+1} 位移到第 {new_index+1} 位，点击「应用变更」保存"
        )
        self._order_dirty = True

        # 重新渲染
        self._render()

        # 让视图滚动到新位置
        self.after(10, self._scroll_to_row, new_index)

    def _scroll_to_row(self, target_row_index):
        """将画布滚动到目标行（在全局 rows 中的索引）大致居中的位置"""
        try:
            items_count = len(self.rows)
            if items_count == 0:
                return

            row_height = 72
            canvas_height = self.canvas.winfo_height()
            total_height = items_count * row_height

            if total_height <= canvas_height:
                return

            target_y = target_row_index * row_height
            target_center_y = target_y + row_height // 2
            ideal_top = target_center_y - canvas_height // 2
            ideal_top = max(0, min(ideal_top, total_height - canvas_height))

            fraction = ideal_top / total_height
            self.canvas.yview_moveto(fraction)
        except Exception:
            pass

    # ======================== 内建预览服务器 ========================
    def _toggle_server(self):
        if self.http_server:
            self._stop_server()
        else:
            self._start_server()

    def _start_server(self):
        try:
            port = 8800
            import socket
            while True:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    if s.connect_ex(('localhost', port)) != 0:
                        break
                    port += 1

            class QuietHandler(SimpleHTTPRequestHandler):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, directory=str(BASE_DIR), **kwargs)
                def log_message(self, format, *args):
                    pass

            self.http_server = HTTPServer(('localhost', port), QuietHandler)
            self.http_server_thread = threading.Thread(
                target=self.http_server.serve_forever, daemon=True)
            self.http_server_thread.start()

            url = f"http://localhost:{port}/index.html"
            webbrowser.open(url)
            self.preview_btn.config(text="🛑 停止服务器", bg=COLOR_DANGER)
            self.status_var.set(f"预览服务器运行中：{url}")
        except Exception as e:
            messagebox.showerror("启动失败", f"无法启动预览服务器：{e}")

    def _stop_server(self):
        if self.http_server:
            self.http_server.shutdown()
            self.http_server = None
            self.http_server_thread = None
            self.preview_btn.config(text="🖥️ 预览前端", bg=COLOR_OK)
            self.status_var.set("预览服务器已停止")

    # ---------- 工具菜单：批量操作 ----------
    def _batch_regen_missing_thumbs(self):
        if not messagebox.askyesno("确认",
                                    "扫描所有原图，给缺失缩略图的生成一份。继续？"):
            return
        ok, fail = 0, []
        for cat in self.rows:
            folder = CLASSIFIED_DIR / f"{cat['id']} {cat['name']}"
            if not folder.is_dir():
                continue
            pic = cat["pic_name"]
            for f in folder.iterdir():
                m = re.match(rf"^{re.escape(pic)}_(\d{{2}})\.jpg$", f.name, re.IGNORECASE)
                if not m:
                    continue
                seq = int(m.group(1))
                thumb = folder / f"{pic}_{seq:02d}_thumb.jpg"
                if thumb.is_file():
                    continue
                try:
                    auto_thumbnail(f, thumb)
                    ok += 1
                except Exception as e:
                    fail.append(f"{f.name}: {e}")
        msg = f"生成成功 {ok} 张。"
        if fail:
            msg += "\n\n失败：\n" + "\n".join(fail[:10])
        messagebox.showinfo("完成", msg)
        self._reload_from_excel()

    def _batch_regen_all_thumbs(self):
        if not messagebox.askyesno(
                "谨慎",
                "会重裁所有原图的缩略图（居中裁剪 60%），\n"
                "覆盖你之前用 thumb_maker 手动裁过的缩略图。\n确认继续？"):
            return
        ok, fail = 0, []
        for cat in self.rows:
            folder = CLASSIFIED_DIR / f"{cat['id']} {cat['name']}"
            if not folder.is_dir():
                continue
            pic = cat["pic_name"]
            for f in folder.iterdir():
                m = re.match(rf"^{re.escape(pic)}_(\d{{2}})\.jpg$", f.name, re.IGNORECASE)
                if not m:
                    continue
                seq = int(m.group(1))
                thumb = folder / f"{pic}_{seq:02d}_thumb.jpg"
                try:
                    auto_thumbnail(f, thumb)
                    ok += 1
                except Exception as e:
                    fail.append(f"{f.name}: {e}")
        msg = f"重生成 {ok} 张。"
        if fail:
            msg += "\n\n失败：\n" + "\n".join(fail[:10])
        messagebox.showinfo("完成", msg)
        self._reload_from_excel()

    def _check_consistency(self):
        try:
            count, warns = regenerate_cats_json(self.rows, CLASSIFIED_DIR, JSON_PATH)
            if not warns:
                messagebox.showinfo(
                    "一致性检查",
                    f"检查通过，共 {count} 只猫咪，未发现问题。\n\n"
                    "cats.json 已同步。")
            else:
                msg = f"警告 {len(warns)} 条：\n\n" + "\n".join("  - " + w for w in warns[:30])
                if len(warns) > 30:
                    msg += f"\n  …以及其他 {len(warns) - 30} 条"
                messagebox.showwarning("一致性检查", msg)
        except Exception as e:
            messagebox.showerror("检查失败", f"{e}\n\n{traceback.format_exc()}")


# ============================================================
# 入口
# ============================================================
def main():
    if not EXCEL_PATH.exists():
        print(f"找不到 {EXCEL_PATH}")
        sys.exit(1)
    if not CLASSIFIED_DIR.exists():
        CLASSIFIED_DIR.mkdir(parents=True, exist_ok=True)
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()