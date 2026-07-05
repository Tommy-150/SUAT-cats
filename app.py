import os
import sys
import json
import shutil
import re
import glob
import threading
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler

import webview
from openpyxl import load_workbook
from PIL import Image

# ================= 路径处理 =================
def get_base_dir():
    if getattr(sys, 'frozen', False):
        # exe 同目录：可写数据（classified/、Excel、JSON 等）
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent

def get_resource_dir():
    if getattr(sys, 'frozen', False):
        # 打包资源目录（只读）：HTML、JS 等
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent

BASE_DIR = get_base_dir()
EXCEL_PATH = BASE_DIR / "统计信息.xlsx"
TAG_COLOR_PATH = BASE_DIR / "tag_color.json"
BAN_WORDS_PATH = BASE_DIR / "ban_words.json"
CATS_JSON_PATH = BASE_DIR / "cats.json"
CLASSIFIED_DIR = BASE_DIR / "classified"

# ================= Excel 读写 =================
class ExcelStore:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb.active
        self.col_idx = {}
        headers = {}
        for c in range(1, self.ws.max_column + 1):
            v = self.ws.cell(row=1, column=c).value
            if v:
                headers[str(v).strip()] = c
        mapping = {
            "id": "编号", "name": "姓名", "gender": "性别",
            "affection": "亲人指数", "status": "状态", "desc": "概要",
            "story": "故事", "pic": "图名"
        }
        for key, kw in mapping.items():
            for h, c in headers.items():
                if kw in h:
                    self.col_idx[key] = c
                    break

    def all_rows(self):
        rows = []
        for r in range(2, self.ws.max_row + 1):
            raw_id = self.ws.cell(row=r, column=self.col_idx["id"]).value
            if raw_id is None or str(raw_id).strip() == "":
                continue
            rows.append({
                "_row": r,
                "id": self._fmt_id(raw_id),
                "name": self._cell_str(r, "name"),
                "gender": self._cell_str(r, "gender") or "unknown",
                "affection": self._cell_int(r, "affection", 1),
                "status": self._cell_str(r, "status") or "normal",
                "desc": self._cell_str(r, "desc"),
                "story": self._cell_str(r, "story"),
                "pic_name": self._cell_str(r, "pic"),
            })
        return rows

    @staticmethod
    def _fmt_id(raw):
        if isinstance(raw, (int, float)):
            return f"{int(raw):02d}"
        return str(raw).strip().zfill(2)

    def _cell_str(self, r, key):
        v = self.ws.cell(row=r, column=self.col_idx[key]).value
        return "" if v is None else str(v).strip()

    def _cell_int(self, r, key, default=1):
        v = self.ws.cell(row=r, column=self.col_idx[key]).value
        if v is None or str(v).strip() == "":
            return default
        try:
            return int(v)
        except:
            return default

    def find_row_by_id(self, cat_id):
        for r in range(2, self.ws.max_row + 1):
            raw = self.ws.cell(row=r, column=self.col_idx["id"]).value
            if raw is None:
                continue
            if self._fmt_id(raw) == cat_id:
                return r
        return None

    def update_row(self, row_index, fields):
        for key, val in fields.items():
            if key in self.col_idx:
                self.ws.cell(row=row_index, column=self.col_idx[key]).value = val

    def append_row(self, fields):
        new_r = self.ws.max_row + 1
        while self.ws.cell(row=new_r, column=self.col_idx["id"]).value not in (None, ""):
            new_r += 1
        self.update_row(new_r, fields)
        return new_r

    def delete_row(self, row_index):
        self.ws.delete_rows(row_index, 1)

    def rewrite_rows(self, rows):
        while self.ws.max_row > 1:
            self.ws.delete_rows(2)
        for idx, cat in enumerate(rows, start=2):
            self.ws.cell(row=idx, column=self.col_idx["id"]).value = int(cat["id"])
            self.ws.cell(row=idx, column=self.col_idx["name"]).value = cat["name"]
            self.ws.cell(row=idx, column=self.col_idx["gender"]).value = cat["gender"]
            self.ws.cell(row=idx, column=self.col_idx["affection"]).value = cat["affection"]
            self.ws.cell(row=idx, column=self.col_idx["status"]).value = cat["status"]
            self.ws.cell(row=idx, column=self.col_idx["desc"]).value = cat["desc"]
            self.ws.cell(row=idx, column=self.col_idx["story"]).value = cat["story"]
            self.ws.cell(row=idx, column=self.col_idx["pic"]).value = cat["pic_name"]

    def next_id(self):
        max_id = 0
        for r in range(2, self.ws.max_row + 1):
            raw = self.ws.cell(row=r, column=self.col_idx["id"]).value
            if raw is None:
                continue
            try:
                max_id = max(max_id, int(str(raw).strip()))
            except:
                pass
        return f"{max_id + 1:02d}"

    def save(self):
        self.wb.save(self.path)

# ================= 图片处理 =================
def auto_thumbnail(src_path, dst_path, size=300):
    img = Image.open(src_path).convert("RGB")
    w, h = img.size
    crop = max(100, int(min(w, h) * 0.6))
    left = (w - crop) // 2
    top = (h - crop) // 2
    img = img.crop((left, top, left + crop, top + crop))
    img = img.resize((size, size), Image.LANCZOS)
    img.save(dst_path, "JPEG", quality=90, optimize=True)

def next_seq_for(folder, pic_name):
    if not folder.is_dir():
        return 1
    used = set()
    for f in folder.iterdir():
        m = re.match(rf"^{re.escape(pic_name)}_(\d{{2}})(?:_thumb)?\.jpe?g$", f.name, re.IGNORECASE)
        if m:
            used.add(int(m.group(1)))
    n = 1
    while n in used:
        n += 1
    return n

# ================= API 类 =================
class ManagerAPI:
    def __init__(self):
        self._store = ExcelStore(EXCEL_PATH)
        self._tag_colors = self._load_json(TAG_COLOR_PATH)
        self._ban_words = self._load_json(BAN_WORDS_PATH).get("words", [])
        self._cat_tags_map = {}
        self._load_cats_json()
        self._ensure_dir()

    def _load_cats_json(self):
        if CATS_JSON_PATH.exists():
            with open(CATS_JSON_PATH, 'r', encoding='utf-8') as f:
                cats = json.load(f)
            for cat in cats:
                self._cat_tags_map[cat['id']] = cat.get('tags', [])

    def _save_cats_json(self):
        if not CATS_JSON_PATH.exists():
            return
        with open(CATS_JSON_PATH, 'r', encoding='utf-8') as f:
            cats = json.load(f)
        for cat in cats:
            cat['tags'] = self._cat_tags_map.get(cat['id'], [])
        with open(CATS_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(cats, f, ensure_ascii=False, indent=2)

    def _cleanup_temp(self):
        temp_dir = BASE_DIR / ".tmp_thumbs"
        if temp_dir.is_dir():
            for f in temp_dir.iterdir():
                if f.is_file():
                    try:
                        f.unlink()
                    except Exception:
                        pass

    def _ensure_dir(self):
        CLASSIFIED_DIR.mkdir(parents=True, exist_ok=True)

    def _load_json(self, path):
        if not path.exists():
            return {}
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _save_json(self, path, data):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # ---------- 公开 API ----------
    def get_cats(self):
        cats = []
        rows = self._store.all_rows()
        for row in rows:
            cat_id = row["id"]
            name = row["name"]
            pic = row["pic_name"]
            folder = CLASSIFIED_DIR / f"{cat_id} {name}"
            avatar = ""
            avatar_hd = ""
            if (folder / f"{pic}_01_thumb.jpg").is_file():
                avatar = f"classified/{cat_id} {name}/{pic}_01_thumb.jpg"
            if (folder / f"{pic}_01.jpg").is_file():
                avatar_hd = f"classified/{cat_id} {name}/{pic}_01.jpg"
            other_photos = []
            pattern = str(folder / f"{pic}_[0-9][0-9].jpg")
            for fpath in sorted(glob.glob(pattern)):
                m = re.match(rf"^{re.escape(pic)}_(\d{{2}})\.jpg$", os.path.basename(fpath))
                if not m:
                    continue
                seq = int(m.group(1))
                if seq < 2:
                    continue
                seq_str = f"{seq:02d}"
                thumb_file = folder / f"{pic}_{seq_str}_thumb.jpg"
                other_photos.append({
                    "thumb": f"classified/{cat_id} {name}/{pic}_{seq_str}_thumb.jpg",
                    "hd": f"classified/{cat_id} {name}/{pic}_{seq_str}.jpg",
                    "seq": seq,
                })
            tags = self._cat_tags_map.get(cat_id, [])
            cats.append({
                "id": cat_id,
                "name": name,
                "gender": row["gender"],
                "avatar": avatar,
                "avatar_hd": avatar_hd,
                "affection": row["affection"],
                "status": row["status"],
                "desc": row["desc"],
                "story": row["story"],
                "tags": tags,
                "otherPhotos": other_photos,
                "pic_name": pic,
            })
        return cats

    def save_cat(self, form):
        cat_id = str(form.get("id", "")).strip().zfill(2)
        name = form.get("name", "").strip()
        pic_name = form.get("pic_name", "").strip()
        if not name or not pic_name:
            return {"success": False, "message": "姓名和图名不能为空"}
        if not re.match(r"^[A-Za-z0-9_]+$", pic_name):
            return {"success": False, "message": "图名只能含英文/数字/下划线"}
        fields = {
            "name": name,
            "gender": form.get("gender", "unknown"),
            "affection": int(form.get("affection", 1)),
            "status": form.get("status", "normal"),
            "desc": form.get("desc", ""),
            "story": form.get("story", ""),
            "pic": pic_name,
        }
        row_idx = self._store.find_row_by_id(cat_id)
        if row_idx:
            # 获取旧数据
            old_rows = self._store.all_rows()
            old_name = ""
            old_pic = ""
            for r in old_rows:
                if r["id"] == cat_id:
                    old_name = r.get("name", "")
                    old_pic = r.get("pic_name", "")
                    break
            # 检测图名变化 → 重命名图片文件
            if old_pic and old_pic != pic_name:
                old_folder = CLASSIFIED_DIR / f"{cat_id} {old_name}"
                if old_folder.is_dir():
                    rename_ops = []
                    try:
                        for f in old_folder.iterdir():
                            m = re.match(rf"^{re.escape(old_pic)}_(\d{{2}})(?:_thumb)?\.jpg$", f.name, re.IGNORECASE)
                            if m:
                                seq = m.group(1)
                                is_thumb = "_thumb" in f.name
                                new_file = f"{pic_name}_{seq}{'_thumb' if is_thumb else ''}.jpg"
                                rename_ops.append((f, old_folder / new_file))
                        for old_path, new_path in rename_ops:
                            old_path.rename(new_path)
                    except Exception as e:
                        for old_path, new_path in reversed(rename_ops):
                            if new_path.exists():
                                try:
                                    new_path.rename(old_path)
                                except Exception:
                                    pass
                        return {"success": False, "message": f"图名重命名失败: {e}"}
            # 检测名字变化 → 重命名文件夹
            if old_name and old_name != name:
                old_folder = CLASSIFIED_DIR / f"{cat_id} {old_name}"
                new_folder = CLASSIFIED_DIR / f"{cat_id} {name}"
                if new_folder.exists():
                    return {"success": False, "message": f"目标文件夹已存在: {new_folder.name}"}
                if old_folder.is_dir():
                    old_folder.rename(new_folder)
            self._store.update_row(row_idx, fields)
        else:
            if not cat_id or cat_id == '00':
                cat_id = self._store.next_id()
                fields["id"] = int(cat_id)
            self._store.append_row({**fields, "id": int(cat_id)})
            folder = CLASSIFIED_DIR / f"{cat_id} {name}"
            folder.mkdir(parents=True, exist_ok=True)
        tags = form.get("tags", [])
        self._cat_tags_map[cat_id] = tags
        # 完整重建 cats.json
        self._regenerate_json()
        self._store.save()
        self._cleanup_temp()
        return {"success": True}

    def delete_cat(self, cat_id):
        cat_id = str(cat_id).strip().zfill(2)
        row_idx = self._store.find_row_by_id(cat_id)
        if row_idx:
            self._store.delete_row(row_idx)
            self._store.save()
        for folder in CLASSIFIED_DIR.glob(f"{cat_id} *"):
            shutil.rmtree(folder, ignore_errors=True)
        if cat_id in self._cat_tags_map:
            del self._cat_tags_map[cat_id]
        self._regenerate_json()
        self._cleanup_temp()
        return {"success": True}

    def upload_image(self, source_path, cat_id, name, pic_name, seq=None):
        cat_id = str(cat_id).strip().zfill(2)
        folder = CLASSIFIED_DIR / f"{cat_id} {name}"
        folder.mkdir(parents=True, exist_ok=True)
        if seq is None:
            seq = next_seq_for(folder, pic_name)
        dst = folder / f"{pic_name}_{seq:02d}.jpg"
        Image.open(source_path).convert("RGB").save(dst, "JPEG", quality=95)
        thumb_dst = folder / f"{pic_name}_{seq:02d}_thumb.jpg"
        auto_thumbnail(dst, thumb_dst)
        self._cleanup_temp()
        return {"success": True, "seq": seq, "thumb": f"classified/{cat_id} {name}/{pic_name}_{seq:02d}_thumb.jpg"}

    def open_file_dialog(self):
        result = webview.windows[0].create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=['Image Files (*.jpg;*.jpeg;*.png)'])
        return result[0] if result else None

    def get_tag_colors(self):
        return self._tag_colors

    def save_tag_colors(self, data):
        # 检测重命名的标签，传播到 cat_tags_map
        old_keys = set(self._tag_colors.keys())
        new_keys = set(data.keys())
        removed = old_keys - new_keys
        added = new_keys - old_keys
        # 如果 key 数量相同但内容不同，可能是重命名
        if len(removed) == 1 and len(added) == 1:
            old_tag = removed.pop()
            new_tag = added.pop()
            for cat_id in list(self._cat_tags_map.keys()):
                if old_tag in self._cat_tags_map[cat_id]:
                    self._cat_tags_map[cat_id] = [
                        new_tag if t == old_tag else t
                        for t in self._cat_tags_map[cat_id]
                    ]
        self._tag_colors = data
        self._save_json(TAG_COLOR_PATH, data)
        self._regenerate_json()
        return {"success": True}

    def get_ban_words(self):
        return {"words": self._ban_words}

    def save_ban_words(self, data):
        self._ban_words = data.get("words", [])
        self._save_json(BAN_WORDS_PATH, {"words": self._ban_words})
        return {"success": True}

    def get_readme(self):
        readme_path = BASE_DIR / "README.md"
        if readme_path.exists():
            return readme_path.read_text(encoding="utf-8")
        return ""

    def delete_photo(self, cat_id, pic_name, seq):
        cat_id = str(cat_id).strip().zfill(2)
        for folder in CLASSIFIED_DIR.glob(f"{cat_id} *"):
            base_name = f"{pic_name}_{int(seq):02d}"
            for ext in ['.jpg', '_thumb.jpg']:
                file_path = folder / f"{base_name}{ext}"
                if file_path.exists():
                    file_path.unlink()
            self._cleanup_temp()
            return {"success": True}
        return {"success": False, "message": "未找到照片"}

    # ===== 裁剪缩略图 =====
    def crop_thumbnail(self, cat_id, pic_name, seq, crop_box):
        cat_id = str(cat_id).strip().zfill(2)
        for folder in CLASSIFIED_DIR.glob(f"{cat_id} *"):
            original_path = folder / f"{pic_name}_{int(seq):02d}.jpg"
            thumb_path = folder / f"{pic_name}_{int(seq):02d}_thumb.jpg"
            if not original_path.is_file():
                return {"success": False, "message": "原图不存在"}
            try:
                left, top, right, bottom = map(int, crop_box)
                img = Image.open(original_path).convert("RGB")
                left = max(0, left); top = max(0, top)
                right = min(img.width, right); bottom = min(img.height, bottom)
                cropped = img.crop((left, top, right, bottom))
                cropped = cropped.resize((300, 300), Image.LANCZOS)
                cropped.save(thumb_path, "JPEG", quality=90, optimize=True)
                self._cleanup_temp()
                return {"success": True}
            except Exception as e:
                return {"success": False, "message": str(e)}
        return {"success": False, "message": "未找到文件夹"}

    # ===== 新增：应用排序（重新编号）=====
    def apply_sort_order(self, ordered_ids):
        rows = self._store.all_rows()
        if len(ordered_ids) != len(rows):
            return {"success": False, "message": "排序列表与数据库不匹配"}
        # 旧 ID → tags 映射
        id_to_tags = {}
        for r in rows:
            id_to_tags[r["id"]] = self._cat_tags_map.get(r["id"], [])
        # 按 ordered_ids 重新排列
        id_to_row = {r["id"]: r for r in rows}
        ordered_rows = [id_to_row[cid] for cid in ordered_ids if cid in id_to_row]
        # 重新编号 + 文件夹重命名
        rename_ops = []
        new_tag_map = {}
        for idx, row in enumerate(ordered_rows, start=1):
            old_id = row["id"]
            new_id = f"{idx:02d}"
            row["id"] = new_id
            new_tag_map[new_id] = id_to_tags.get(old_id, [])
            if old_id != new_id:
                old_folder = CLASSIFIED_DIR / f"{old_id} {row['name']}"
                new_folder = CLASSIFIED_DIR / f"{new_id} {row['name']}"
                if old_folder.is_dir() and not new_folder.exists():
                    rename_ops.append((old_folder, new_folder))
        for old_f, new_f in rename_ops:
            old_f.rename(new_f)
        self._cat_tags_map = new_tag_map
        self._store.rewrite_rows(ordered_rows)
        self._store.save()
        self._regenerate_json()
        self._cleanup_temp()
        return {"success": True, "count": len(ordered_rows)}

    # ===== 新增：独立应用标签 =====
    def apply_tags(self, cat_id, tags):
        cat_id = str(cat_id).strip().zfill(2)
        self._cat_tags_map[cat_id] = tags
        self._regenerate_json()
        return {"success": True}

    # ===== 新增：保存编者留言 =====
    def save_readme(self, content):
        readme_path = BASE_DIR / "README.md"
        try:
            readme_path.write_text(content, encoding="utf-8")
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

    # ===== 新增：替换主图（文件交换）=====
    def swap_to_main(self, cat_id, pic_name, seq):
        cat_id = str(cat_id).strip().zfill(2)
        target_seq = int(seq)
        if target_seq == 1:
            return {"success": False, "message": "已经是主图"}
        for folder in CLASSIFIED_DIR.glob(f"{cat_id} *"):
            try:
                def swap_pair(a, b):
                    for suffix in [".jpg", "_thumb.jpg"]:
                        a_file = folder / f"{pic_name}_{a:02d}{suffix}"
                        b_file = folder / f"{pic_name}_{b:02d}{suffix}"
                        tmp_file = folder / f"{pic_name}_{a:02d}{suffix}.tmpswap"
                        if a_file.is_file():
                            a_file.rename(tmp_file)
                        if b_file.is_file():
                            b_file.rename(a_file)
                        if tmp_file.is_file():
                            tmp_file.rename(b_file)
                swap_pair(1, target_seq)
                self._cleanup_temp()
                return {"success": True}
            except Exception as e:
                return {"success": False, "message": str(e)}
        return {"success": False, "message": "未找到文件夹"}

    def open_folder(self, cat_id, name):
        cat_id = str(cat_id).strip().zfill(2)
        import subprocess
        folder = CLASSIFIED_DIR / f"{cat_id} {name}"
        if folder.is_dir():
            subprocess.Popen(['explorer', str(folder)])
            return {"success": True}
        return {"success": False, "message": "文件夹不存在"}

    # ===== 新增：行顺序移动 =====
    def move_row(self, cat_id, direction, step=1):
        cat_id = str(cat_id).strip().zfill(2)
        step = int(step)
        rows = self._store.all_rows()
        idx = next((i for i, r in enumerate(rows) if r["id"] == cat_id), None)
        if idx is None:
            return {"success": False, "message": "未找到该猫咪"}
        new_idx = max(0, idx - step) if direction == "up" else min(len(rows) - 1, idx + step)
        if new_idx == idx:
            return {"success": True, "message": "已在边界"}
        rows.insert(new_idx, rows.pop(idx))
        # 重建 Excel
        self._store.rewrite_rows(rows)
        self._store.save()
        self._regenerate_json()
        return {"success": True}

    # ===== 新增：多文件选择对话框 =====
    def open_file_dialog_multi(self):
        result = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=True,
            file_types=['Image Files (*.jpg;*.jpeg;*.png)']
        )
        return list(result) if result else []

    # ===== 新增：批量上传图片 =====
    def upload_images(self, source_paths, cat_id, name, pic_name):
        results = []
        for src in source_paths:
            r = self.upload_image(src, cat_id, name, pic_name)
            results.append(r)
        return {"success": True, "results": results}

    # ===== 新增：重命名图名（修改所有文件）=====
    def rename_pic_name(self, cat_id, name, old_pic, new_pic):
        cat_id = str(cat_id).strip().zfill(2)
        if not re.match(r"^[A-Za-z0-9_]+$", new_pic):
            return {"success": False, "message": "图名只能含英文/数字/下划线"}
        for folder in CLASSIFIED_DIR.glob(f"{cat_id} *"):
            if not folder.is_dir():
                continue
            try:
                for f in sorted(folder.iterdir()):
                    m = re.match(rf"^{re.escape(old_pic)}_(\d{{2}})(_thumb)?\.(jpe?g)$", f.name, re.IGNORECASE)
                    if m:
                        seq = m.group(1)
                        suffix = m.group(2) or ""
                        ext = m.group(3)
                        new_name = f"{new_pic}_{seq}{suffix}.{ext}"
                        f.rename(folder / new_name)
                return {"success": True}
            except Exception as e:
                return {"success": False, "message": str(e)}
        return {"success": False, "message": "未找到文件夹"}

    # ===== 新增：完整重建 cats.json =====
    def regenerate_cats_json(self):
        return self._regenerate_json()

    def _regenerate_json(self):
        warnings = []
        cats = []
        rows = self._store.all_rows()
        for row in rows:
            cat_id = row["id"]
            name = row["name"]
            pic = row["pic_name"]
            folder = CLASSIFIED_DIR / f"{cat_id} {name}"
            folder_json = f"classified/{cat_id} {name}"
            if not folder.is_dir():
                warnings.append(f"[{cat_id}] 文件夹缺失: {cat_id} {name}")
                continue
            if not pic:
                warnings.append(f"[{cat_id}] 图名为空")
                continue
            avatar = f"{folder_json}/{pic}_01_thumb.jpg"
            avatar_hd = f"{folder_json}/{pic}_01.jpg"
            if not (folder / f"{pic}_01.jpg").is_file():
                warnings.append(f"[{cat_id}] 缺头像原图 {pic}_01.jpg")
            if not (folder / f"{pic}_01_thumb.jpg").is_file():
                warnings.append(f"[{cat_id}] 缺头像缩略图 {pic}_01_thumb.jpg")
            other_photos = []
            pattern = str(folder / f"{pic}_[0-9][0-9].jpg")
            for fpath in sorted(glob.glob(pattern)):
                m = re.match(rf"^{re.escape(pic)}_(\d{{2}})\.jpg$", os.path.basename(fpath))
                if not m:
                    continue
                seq = int(m.group(1))
                if seq < 2:
                    continue
                seq_str = f"{seq:02d}"
                thumb_file = folder / f"{pic}_{seq_str}_thumb.jpg"
                if not thumb_file.is_file():
                    warnings.append(f"[{cat_id}] 缺缩略图 {thumb_file.name}")
                other_photos.append({
                    "thumb": f"{folder_json}/{pic}_{seq_str}_thumb.jpg",
                    "hd": f"{folder_json}/{pic}_{seq_str}.jpg",
                })
            tags = self._cat_tags_map.get(cat_id, [])
            cats.append({
                "id": cat_id, "name": name,
                "gender": row["gender"] or "unknown",
                "avatar": avatar, "avatar_hd": avatar_hd,
                "affection": row["affection"],
                "status": row["status"] or "normal",
                "desc": row["desc"], "story": row["story"],
                "tags": tags,
                "otherPhotos": other_photos,
            })
        with open(CATS_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(cats, f, ensure_ascii=False, indent=2)
        return {"success": True, "count": len(cats), "warnings": warnings}

    # ===== 新增：获取照片健康状态 =====
    def get_photo_health(self, cat_id, pic_name):
        cat_id = str(cat_id).strip().zfill(2)
        for folder in CLASSIFIED_DIR.glob(f"{cat_id} *"):
            issues = []
            for f in sorted(folder.iterdir()):
                m = re.match(rf"^{re.escape(pic_name)}_(\d{{2}})\.jpg$", f.name, re.IGNORECASE)
                if not m:
                    continue
                seq = int(m.group(1))
                thumb_file = folder / f"{pic_name}_{seq:02d}_thumb.jpg"
                if not thumb_file.is_file():
                    issues.append({"seq": seq, "issue": "missing_thumb", "file": f"{pic_name}_{seq:02d}_thumb.jpg"})
            return {"success": True, "issues": issues}
        return {"success": False, "message": "未找到文件夹"}

    # ===== 新增：清理临时文件 =====
    def cleanup_temp_files(self):
        temp_dir = BASE_DIR / ".tmp_thumbs"
        deleted = []
        size_freed = 0
        if temp_dir.is_dir():
            for f in temp_dir.iterdir():
                if f.is_file():
                    try:
                        size_freed += f.stat().st_size
                        f.unlink()
                        deleted.append(f.name)
                    except Exception:
                        pass
        return {"success": True, "deleted": len(deleted), "size_freed": size_freed, "files": deleted[:50]}

    # ===== 新增：扫描孤立文件 =====
    def scan_orphan_files(self):
        orphans = []
        if not CLASSIFIED_DIR.is_dir():
            return {"success": True, "orphans": []}
        # 收集所有合法的图片文件名模式
        rows = self._store.all_rows()
        valid_patterns = set()
        for row in rows:
            pic = row["pic_name"]
            if pic:
                valid_patterns.add(pic)
        # 扫描 classified 下所有文件夹
        for folder in sorted(CLASSIFIED_DIR.iterdir()):
            if not folder.is_dir():
                continue
            for f in sorted(folder.iterdir()):
                if not f.is_file():
                    continue
                name = f.name
                # 检查是否为有效猫图片（匹配任一 pic_name_xx.jpg 格式）
                matched = False
                for pic in valid_patterns:
                    if re.match(rf"^{re.escape(pic)}_\d{{2}}(_thumb)?\.jpe?g$", name, re.IGNORECASE):
                        matched = True
                        break
                    # 也匹配 tmpswap 文件
                    if name.endswith('.tmpswap'):
                        matched = True
                        break
                if not matched:
                    orphans.append(str(f.relative_to(CLASSIFIED_DIR)))
        return {"success": True, "orphans": orphans}

    # ===== 新增：删除孤立文件 =====
    def delete_orphan_files(self, paths):
        deleted = []
        for rel_path in paths:
            full_path = CLASSIFIED_DIR / rel_path
            if full_path.is_file():
                try:
                    full_path.unlink()
                    deleted.append(rel_path)
                except Exception:
                    pass
        return {"success": True, "deleted": deleted}

# ================= HTTP 服务器 =================
def start_http_server():
    base_dir = get_base_dir()
    # 确保必要目录存在
    (base_dir / 'classified').mkdir(parents=True, exist_ok=True)
    
    class Handler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(base_dir), **kwargs)
        def log_message(self, format, *args):
            pass
    server = HTTPServer(('127.0.0.1', 0), Handler)
    port = server.server_address[1]
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return f'http://127.0.0.1:{port}/app_manager.html', server

# ================= 主入口 =================
if __name__ == '__main__':
    api = ManagerAPI()
    url, server = start_http_server()
    window = webview.create_window(
        title='🐱 SUAT 猫咪管理器',
        url=url,
        js_api=api,
        width=1200,
        height=800,
        min_size=(900, 600)
    )
    webview.start()